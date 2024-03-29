import re

from PIL import ImageEnhance

VER = [
        "ver_20240329",
        "ver_20240325",
        "ver_20240324",
        "ver_20240320",
        "ver_20240317",
        "ver_20240313",
        "ver_20240312",
        "ver_20240311",
        "ver_20240301",
        "ver_20240230"
    ]

DEV_MODE = False         # Production == False, Development == True
LOCAL_HC_FLAG = False
OCR_FLAG = 999             # {1: Capture screenshot & OCR analysis, 2: Manually open web and copy empNum, 3: Reading all.txt file, else: Open pup up for email and open url - nneds to be connected to HP network}
# TODO: Insert high level explanation of the app.
"""

"""
import json
import csv
import numpy as np
import time
import os
import pandas as pd
import subprocess
import platform
from dateutil.parser import parse
from datetime import datetime, timedelta, date
from threading import Thread
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog, messagebox
import hashlib
from pathlib import Path
#   -------------------------------------------------------------------------------------------------------------      #
#   -------------------------------------------------------------------------------------------------------------      #
#   -------------------------------------------------------------------------------------------------------------      #
today_date          = datetime.now().strftime("%Y-%m-%d")
ORG_CWD             = Path.cwd()                                                                                        # Save the current working directory
HC_TARGET_FOLDER    = Path(r'C:\Users\biton\OneDrive - HP Inc\Core Team - General\PMO - Khalil and Efrat\HC')
LIMOR_FOLDER        = Path(r'C:\Users\biton\OneDrive - HP Inc\Core Team - General\Employees Experience - Limor\HC vs. System discrepancies check')
HC_SHEET_NAME       = "Employee Headcount"
HC_COL              = [
                        "Workstream",
                        "Drupa role (booth duty) / Visitor / Channel partner",
                        "Manager name",
                        "Last name",
                        "First name",
                        "Employee Email",
                        "Employee direct manager email",
                        "HP Employee#",
                        "Arriving from",
                        "Airport/Country",
                        "Comments",
                        'hc_per_tup', '1st_hc', 'hc_1st_hotel_in', 'hc_1st_hotel_out', '2nd_hc', 'hc_2nd_hotel_in', 'hc_2nd_hotel_out', '3rd_hc', 'hc_3rd_hotel_in', 'hc_3rd_hotel_out', 'source_HC',
                        ]
CANAP_COL           = [
            'firstName', 'lastName', 'jobTitle', 'company', 'country', 'Organization', 'email', 'phone_number',
            'type', 'ticketName', 'status', 'OtherSocialMediaURL', 'linkedInProfile', 'telNum', 'twitter', 'hotel',
            'hotelCheckIn', 'hotelCheckOut', 'Documents', 'marketingUpdates', 'First seen', 'Login time', 'Last seen',
            'Registered', 'checkin', 'Last updated date', 'Last updated by', 'Last updated attributes', 'nickName',
            'gender', 'dietary', 'Total watch time', 'Total sessions booked', '_views', 'Region', 'Country',
            'APJ - Sub Region & HP Office', 'EMEA - Sub Region & HP Office', 'AMS - Sub Region & HP Office',
            'Languages (You may select more than one)', 'If other, please specify language', 'Are you HP employee?',
            'HP Business', 'HP Business - Industrial', 'Department/Function', 'Country Coverage', 'Business Segment',
            'HP Job Title', 'Role at Show', 'Tshirt Size', 'Manager Name', 'Manager Email Address',
            'I am an expert in', 'Type', 'Company Name', 'Job Title',
            'Will you be attending the sales training day on May 27th?', 'Terms and Conditions',
            'I would like to attend the following dinner events:', 'What’s the Name of Your HP Sales Account Manager?', 'Additional Info', 'Nationality',
            'Passport Number', 'Passport Expiry Date', 'Passport Issue Date', 'Passport Issued by',
            'Date of Birth', 'Dietary Restrictions', 'Life-threatening Allergy',
            'AMS Taste and Talks Event Details', 'Do you need any assistance with regards to accessibility?', 'Emergency Contact Name',
            'Emergency Contact Number', 'Visa Requirement', 'Do you need a visa for your trip?',
            'Accommodation Reminder', 'Do you require accommodation?', 'Reminder',
            'Indicate below if you have any special accommodation requests related to your religion or accessibility requirements. Please note that your request is not guaranteed and will be subject for approval.',
            '1st Hotel Check in', '1st Hotel Check out',
            '2nd Hotel Check in', '2nd Hotel Check out',
            '3rd Hotel Check in', '3rd Hotel Check out',
            'Do you observe Shabbat?', 'Plan your Trip',
            'How will you be travelling to Germany?', 'Flight 1', 'Flight 2', 'Flight 3',
            'Add additional flight? [Flight 2]', 'Add additional flight? [Flight 3]',
            'Sub Region - APJ', 'Sub Region - AMS', 'Sub Region - EMEA', 'Your HP Office Address - EMEA',
            'Flight 1: Arrival Airport', 'Flight 1: Arrival Date', 'Flight 1: Arrival Time', 'Flight 1: Flight Number',
            'Flight 2: Arrival Airport', 'Flight 2: Arrival Date', 'Flight 2: Arrival Time', 'Flight 2: Flight Number',
            'Flight 3: Arrival Airport', 'Flight 3: Arrival Date', 'Flight 3: Arrival Time', 'Flight 3: Flight Number',
            'Your HP Office Address - APJ', 'Your HP Office Address - AMS', 'preferredTimezone', 'experience',
            '1st_cnpi', '2nd_cnpi', '3rd_cnpi', 'cnpi_per_tup', 'db_id', 'source_CP',
            ]
### Registration Analysis   -----------------------------------------------------------------------------------------  #
# REPORT_APPROVED_PPL_FOR_FLIGHTS_COLS = ['firstName', 'lastName', 'country', 'email', 'type',
#                                         'ticketName', 'status', 'Country', 'Role at Show',
#                                         '1st Hotel Check in', '1st Hotel Check out',
#                                         '2nd Hotel Check in', '2nd Hotel Check out',
#                                         '3rd Hotel Check in', '3rd Hotel Check out',
#                                         'How will you be travelling to Germany?']
OPEN_WORKSTREAM = ['Operations', 'Employee Services', 'CoreTeam']
DRUPA_CostCenter_Allocation_Workstreams = ['Operations', 'Employee Services', 'CoreTeam']
CoreTeam_emp_onDrupa_CC = [('Tweig', 'Idan'), ('Thaler', 'Efrat'), ('Khalil', 'Eid')]
tshirts_relevancy_workstreams = ['Operations']
Workstream_Pairs_Mapping = {
    'Operations'            : ['Operations (Tech & Operators)', 'Operations (Tech & Operators) '],
    'Sales'                 : ['Sales Account Manager', 'Sales Manager', 'Sales Director', 'Sales Executive'],
    'Sales Operations'      : ['Sales Operations (WW)'],
    'CoreTeam'              : ['Drupa HP core team', 'drupa HP core team'],
    'Employee Services'     : ['Drupa HP core team', 'drupa HP core team'],
    'Marketing'     	    : ['Marketing'],
    'Customer Services'	    : ['Customer services'],
    'PSB'	                : ['PSB (Indigo product GBU)'],
    'Communication'	        : ['Communication'],
    'PWI'	                : ['PWI'],
    'Scitex'	            : ['Corrugated'],
    'Corrugated'            : ['Corrugated'],
    'GSA'                   : ['GSA'],
    'Service'               : ['Service'],
}
EXCLUDING_FROM_HOTELS_AND_FLIGHTS_EMAILS = [
    'joachim.boessler@hp.com',          # I do NOT need a hotel. I live nearby. Need access to the floor from 13th May to 12th June
    'mark.lemmen@hp.com',               # I do not need accommodation as I life close to DRUPA and will commute by car.
    'philippe.mucher@hp.com',           # I will be driving in every day from home, NO HOTEL NEEDED
    'erik.brammer@hp.com',              # No hotel to be booked for me! Will stay in my camper van. Dates shown here are just for my attendance at drupa.
    'martinus.van.elderen@hp.com'       # No need for hotel only parking spot at Drupa
]
COLS_FOR_LIMOR = [
    'Employee Email', 'First name', 'Last name', 'overall_approval', 'Registered_or_not',
    'Hotels_inout_approval', 'role_flag', 'Workstream', 'Role at Show',
    'Drupa role (booth duty) / Visitor / Channel partner',
    'hc_then_cnpi', '1st_in', '1st_out', '2nd_in', '2nd_out', '3rd_in', '3rd_out',
    'Comments', 'type', 'ticketName', 'gender', 'Country', 'Passport Expiry Date', 'Tshirt Size',
    'Do you observe Shabbat?', 'dietary', 'Life-threatening Allergy', 'Dietary Restrictions',
    'Indicate below if you have any special accommodation requests related to your religion or accessibility requirements. Please note that your request is not guaranteed and will be subject for approval.',
    'hc_per_tup', 'cnpi_per_tup', '1st Hotel Check in', '1st Hotel Check out', '2nd Hotel Check in',
    '2nd Hotel Check out', '3rd Hotel Check in', '3rd Hotel Check out', 'hc_original_index', 'HP Employee#',
    'Last updated date', 'Last updated attributes', 'hotelsExceptions'
]
DRUPA_DATES = {'start': '2024-05-01', 'end': '2024-06-16'}
#   -------------------------------------------------------------------------------------------------------------      #
# TODO: Unique values
1/1
# uniq_HC_Workstream         = ['CoreTeam', 'Marketing', 'Employee Services', 'Customer Services', 'Communication', 'Sales Operations', 'Sales', 'Sustainability', 'PWI', 'Operations', 'PSB', 'LF', 'Budget', 'Scitex', 'GSA', 'Service']
# uniq_CNPI_ticketName       = ['HP Operations', 'Canapii Team - Test Ticket', 'HP Duty Employees']
# uniq_CNPI_type             = ['HP Operations', 'HP Duty Employees']
# uniq_HC_Type               = ['Partner', 'Contractor', 'Customer', 'Channel Partner']
# uniq_HC_Drupa_role_        = ['Drupa 2024 PMO Manager', 'Core Team', 'Employee Service', 'Customer Service', 'Internal Communication', 'Operations Lead', 'Drupa PMO', 'Sales operations (Drakkart, proposals, contracts)', 'Sales  ', 'Sustainability', 'GBU Marketing PWP - Booth Duty', 'Security', 'Logistics', 'Build up and Tear Down + Infrastructure', 'IT', 'Ops Indigo Show', 'PWI Ops', 'EHS', 'Indigo Sustainability', 'Sourcing', 'Supply Chain', 'Demoing the press during the show - Indigo', 'LF Ops', 'NPI Manager', 'LF ISR German', 'LF ISR English and Nordics', 'LF Business Management', 'LF Sales CEE', 'LF Sales NWE', 'Sustainability Pitch', 'LF Sales CEMA', 'LF Sales APJ', 'LF Sales AMS', 'LF PR', 'LF Marketing Management', 'Visitor', 'Ops indigo Show', 'Communication lead for Drupa', 'external communication support', 'external communication German media', 'external communication Large Format', ' external communication TBC', 'Event Manager and Reception Manager', 'Reception', 'Booth Tour', 'Private Demo', 'Meeting Rooms', 'Application Showcase SCITEX', 'Executives', 'Sales Support - Scitex', 'Sales Account Manager PWP', 'PWI R&D Lead', 'EMEA Industrial Sales Operations Manager', 'EMEA Industrila Sales Enablement Manager', 'WW Marketing', 'SoMe & Comms Team (2 designers, 2 videographers, editing)', 'Global Account Business Manager (Oran)', 'Global program Business Manager (Oran)', 'Demoing Workflow SW and Solutions (Indigo)', 'Marketing - Booth duty  ', 'Application Showcase GCP', 'Application Showcase L&P', 'Techfair', 'Security Applications', 'Alliance Managers (Media & Finishing Partners)', 'Sales Support - Indigo', 'Sales Support', 'LF GBU Product ', ' LF Ops Lead', 'LFPro  MV PM', 'LFPro  operator', 'Demoing Service Portfolio', 'Service gaming zone', 'Demoing the press during the show - PWP', 'Service booth manager', 'Sales Account Managers L&P', 'Sales Account Managers GCP', 'Solution Consultant (Indigo & PWP)', 'Regional GM & Staff', 'Sales Account Manager Indirect', 'P&S (Product and Solutions)', 'Sales Support supplies', 'SDR (sales development)', 'GEC (graphic experience center)', 'Sales Account Manager GCP/L&P', 'Regional Business Manager incl. Indirect ', 'Sales Account Manager SCITEX', 'Duty', 'Deal Support', 'Duty supporting Deals', 'Booth duty', 'Management Visit', 'Booth duty/Visitor']
# uniq_HC_Comments           = ['Receptionist/lead capturer/meeting room manager', 'Receptionist/lead capturer', 'Sustainability lead', 'Core Team - PWP GBU Drupa Lead', 'Place holder for all', 'Lead Logistics', 'Lead Build up and Tear Down + Infrastructure', 'Lead IT', 'Lead  Indigo Show', 'Tech Lead', 'Lead EHS', 'EHS #2', 'EHS #3', 'Finishing Lead  & direct2finish', 'Lead Sustainability', 'Sustainability #2', 'Lead Sourcing', 'Lead Supply chain ', 'Tech. Leader', 'Press Operator V12', 'Press Operator V12 + MATP', 'Press Operator 200K (Was Martin)', 'Press Operator 200K', 'Press Operator 200K - MATP', 'Press Operator 120K', 'Press Operator 120K - second shift', 'Temp (Ops and logistics)', 'Tech fair - special inks ', 'nearby Hotel', 'nearby hotel', 'Award winner #1', 'Award winner #2', 'Award winner #3', 'Award winner #4', 'Award winner #5', 'Award winner #6', 'Award winner #7', 'Award winner #8', 'Award winner #9', 'Award winner #10', 'Press Operator 120K + MATP', 'Press Operator 18K', 'Press Operator 35K', 'Press Operator 6K', 'Press Operator 6K (Was Michael Cox)', 'Press Operator 7K + 35K Backup + MATP', 'Press Operator 7K', 'Press Operator 7K Secure', 'Press Operator 18K Demo', 'Press Operator 120K Demo + MATP', 'Press Operator 120K Demo', 'Press Operator + CE 7K Secure', 'DFE Leader+Performer', 'DFE Performer', 'DFE Performer 1st Half', 'DFE Performer 2nd half', 'CE V12 + MATP', 'CE 200K', 'CE 120K 2nd Half +MATP', 'CE 18K 1st Half', 'CE 18K 2nd Half + MATP', '200K DFE Show', 'CE 35K 2nd Half', 'CE 6K + MATP', 'CE 7K', 'Process Expert V12', 'WH Expert + Process', 'PLC V12 - Install + 1st Half', 'PLC Commercial Install + 1st Half', 'PLC Industrial Install + 1st Half', 'SW / QA V12', 'SW / QA 120K Install + 1st Half', 'SW / QA 200K Install + 1st Half', 'SW / QA 2nd Half All platforms', 'SW/QA 35K Instllation', 'SW / QA 18K&35K Install + 1st Half', 'Electronic', 'Automation R&D + show booth duty', 'solutions', 'Studio Pre-press', 'Studio - designer ', 'Studio - production manager', 'Production Industrial', 'Production Commercial', 'Studio logistics', 'Studio Production', 'Build up / Tear down Lead', 'Infrastructure lead', 'installer (A) 120K#1 Area 1', 'installer (B) 120K #1 Area 1', 'installer (C) 120K #2 Area 2', 'installer (B) 6K', 'installer (E) 120K #3 Area 3 - Demo Room', 'installer (F) 120K #3 Area 3 - Demo Room', 'installer (A) 18K #1 Area 1', 'installer (B) 18K #1 Area 1', 'installer (C) 18K #2 Area 2 Demo Room', 'installer (D) 18K#2 Area 2 Demo Room', 'installer (A) 200K', 'installer (B) 200K', 'installer (C) 200K', 'installer (A) 7K ', 'installer (A) 7K Secure', 'installer (B) 7K Secure', 'installer (A) 35K + 35K CE Show 1st Half', 'installer (B) 35K', 'installer (D) 35K', 'installer (C) 35K', 'installer (A) 6K', 'installer (D) 120K #2 Area 2', 'installer (A) V12 ', 'installer (B) V12 ', 'installer (C) V12 ', 'installer (D) V12 ', 'DFE lead', 'DFE commercial', 'Ser 3/4install  lead+expert (6, 7, 200)', '18K/120K install SF lead + CE 120K 1st half', 'Ser. 6 install  lead+expert', 'Writing head expert', 'Web handaling expert (V12 NSE / NSRW + AMR & UMDC', 'PQ / Process Expert', 'Web handaling expert (V12 NSE / NSRW) PH V12 + 200K Slitter', 'SW Specialist all product', 'Tear down 2', 'MATP 120K #2', 'MATP 200K', 'MATP 35K', 'MATP 18K #2', 'MATP 6K ', 'MATP 7K', 'MATP V12', 'Automation R&D', 'A2200 GBU Technical Coordinator', 'Build up A2200 Install project manager', 'Build up A2200 GBU HW Mechanic Eng', 'Build up A2200 DFE Tech', 'Build up A2200 Validation Tech', 'Build up A2200 IC/wiring Tech (First Shift)', 'Build up A2200 IC/wiring Tech (Night Shift)', 'Build up A2200 Vision Tech (Second Shift)', 'Show A2200 PWI solutions Hunkeler support engineer', 'Show A2200 PWI Solutions H&B Support Engineer', 'Show A2200 PWI GBU ME', 'Show A2200 SW support', 'Show A2200 WS', 'Show A2200 Press Tech', 'Show A2200 Press Operator - Demoing the press during the show - PWP', 'KN Manager ', 'KN Manager 2', 'Messe inventory manager (SL) ', 'Messe inventory manager 2 (SL) ', 'SP manager', 'SP manager 3', 'PWI manager', 'PWI manager 2 & finishing', 'External warehouse manager (Media)', 'External warehouse manager 2 (Media) ', 'Large Format', 'corrugated - booth operations', 'corrugated', 'corrugated - technical knowledge support from GBU ', 'Corrugated Mgmt - Products & Solutions Manager', 'Corrugated Product Marketing - Workflow & Partners', 'Corrugated Product Marketing - Prepress & Color mgmt', 'Corrugated Sales Acoount Manager', 'corrugated & commercial', 'corrugated & commercial - Operations Director', 'commercial', 'EMEA Industrial Sales Operations Manager', 'EMEA Industrila Sales Enablement Manager', 'Industrial Sales Operations', 'Commercial Mgmt - Product Management NDA room', 'Commercial Market Development - Segment applications area', 'Commercial Market Development - A2200 Product manager', 'Commercial Print Booth setup Support - customer success and applications area', 'Commercial - Install & Support - Pre-press', 'Commercial Product Marketing - Contiweb, H&B and T485 Product Manager + NDA room', '3rd Party Finishing Program Manager', 'Commercial Product Marketing - Finishing and Workflow', 'Commercial Product Marketing - Sustainability', 'Commercial Product Marketing - Sustainability and Product Support', 'Commercial Product Marketing - Partner Manager', 'PWP Colorpro media launch', 'Commercial Product Marketing - TIJ Technology NDA room', 'Category AMS - demo', 'Category EMEA - demo', 'SA EMEA - expert demos', 'SA EMEA - expert demos - Italian', 'SA EMEA - expert demos - German? - Workflow anf finishing', 'SA EMEA - expert demos - French - workflow and color management - TAG Elite', 'SA AMS Workflow Booth Duty', 'Service Workflow Booth Duty', 'Social Media', 'Cimpress Global Account Manager ', 'Avery Dennison Global Account Manager ', 'ePac Global Account Manager ', 'SealedAir & Amcor Global Account Manager ', 'CCL & Westrock Global Account Manager ', 'Amazon and RRD Global Account Manager ', 'Shutterfly Global Account Manager ', 'MCC Global Account Manager ', 'MCC Global Account operation Manager ', 'All4Labels Global Account Manager', 'Global Head, GSA & ISS Sales', 'Shutterfly Global Account operation Manager ', 'Cimpress Global Program Manager ', 'ePac Global Program Manager ', 'MCC & CCL Global Program Manager ', 'All4Labels Global Program Manager ', 'Global future accounts program manager', 'Avery Dennison Global Program Manager ', 'GSA Business operations manager', 'ISS manager', 'Head of Site Flow\u200b', 'Receptionist/lead capturer/actor', 'Sufa + FC + BP', 'Labels', 'GCP ', 'floater ', 'NDA room/Tech fair - LEP/x', '7K secure ', '7K Secure', '7K Secure ', 'V12', 'Solutions - creativity ', '7K', 'Solutions - floater ', 'all presses', 'Solutions - workflow L&P', 'Partners - finishing ', 'floater - GCP ', 'management ', 'Brand Innovation/TouchPoint Packaging', 'Solution - quality ', 'Media & Materials - L&P ', 'Media & Materials - GCP ', 'floater - L&P', 'Competitive Intelligence', '200K ', 'Inks - GCP ', 'Inks - L&P ', '6K', '35K', 'Sustainability - general station ', 'Sustainability - GCP ', 'Sustainability - L&P ', 'Sufa', 'Amit ', 'Training + General Support', 'FP ', 'solutions ', 'Sufa HD', 'Sufa HD ', 'Sufa demo ', 'Sufa ', 'Amit demo ', '7K ', '6K  ', 'V12 NSW', '200K', 'Solutions - UMDC', 'Solutions - quality GCP ', 'Tech fair - Digital HTL ', 'Tech fair - RFID', 'Strategic partnership ', 'Amit', 'Sufa A', 'Robots ', '6K ', 'V12 ', '35K ', 'Soluions ', 'Brand Protection ', '7K  ', 'Amit Demo ', 'Sufa Demo ', 'PWI proffesional services booth', ' services booth future tools', ' services booth Professional services week2', ' services booth Professional services week1', 'Gaming zone', 'Amit week1', 'Amit week2', '7k week1', '7k week2', 'PWI A220 week1 ', 'PWI A220 week2', 'Sufa HD+ Sufa A week1', 'Sufa HD+ Sufa A week2', 'managing service - core team', '200K week1', '200K week2', 'V12 week1', 'V12 week2', '6k week1', 'Total per day', '6k week2', ' services booth current tools', 'Supplies business managre', 'Supplies Category']
# uniq_CNPI_Role_at_Show     = ['drupa HP core team', 'Drupa HP core team', 'Operations (Tech & Operators)', 'Operations (Tech & Operators) ', 'Other ', 'Booth duty ', 'Sales', 'Booth duty']
# analyzed_reg_df_columns    = [
                            #     'db_id', 'hc_original_index', 'Workstream', 'Drupa role (booth duty) / Visitor / Channel partner',
                            #     'Manager name', 'Last name', 'First name', 'Employee Email', 'Employee direct manager email',
                            #     'HP Employee#', 'Arriving from \nAirport/Country', 'Comments', 'hc_per_tup', '1st_hc', '2nd_hc', '3rd_hc',
                            #     'firstName', 'lastName', 'email', 'type', 'ticketName', 'Role at Show', 'Country', 'cnpi_per_tup',
                            #     '1st_cnpi', '2nd_cnpi', '3rd_cnpi', 'How will you be travelling to Germany?', 'jobTitle', 'company',
                            #     'Organization', 'phone_number', 'Passport Number', 'Passport Expiry Date', 'gender', 'HP Job Title',
                            #     'Tshirt Size', 'Manager Name', 'Manager Email Address', 'Type', 'Company Name', 'Job Title',
                            #     '1st Hotel Check in', '1st Hotel Check out', '2nd Hotel Check in', '2nd Hotel Check out',
                            #     '3rd Hotel Check in', '3rd Hotel Check out', 'Flight 1', 'Flight 2', 'Flight 3',
                            #     'Add additional flight? [Flight 2]', 'Add additional flight? [Flight 3]', 'Flight 1: Arrival Airport',
                            #     'Flight 1: Arrival Date', 'Flight 1: Arrival Time', 'Flight 1: Flight Number', 'Flight 2: Arrival Airport',
                            #     'Flight 2: Arrival Date', 'Flight 2: Arrival Time', 'Flight 2: Flight Number', 'Flight 3: Arrival Airport',
                            #     'Flight 3: Arrival Date', 'Flight 3: Arrival Time', 'Flight 3: Flight Number'
                            #     ]
1/1
#   -------------------------------------------------------------------------------------------------------------      #
def create_identifier(row, columns):
    concatenated = ''.join([(str(row[column]).lower() if row[column] is not None else 'none') for column in columns])   # Concatenate the lowercased strings of specified columns, replace None with 'none'
    identifier = hashlib.sha256(concatenated.encode()).hexdigest()                                                      # Generate a hash as the identifier - using SHA256 here for example
    return identifier
def searching_tuple_and_list_of_tuples(target_tuple, list_of_tuples) -> bool:
    list_of_tuples = [tuple(sorted((name.lower() for name in tup))) for tup in list_of_tuples]                          # temp_coreTeam_emp_on_drupa = [tuple(sorted((name.lower() for name in tup))) for tup in CoreTeam_emp_onDrupa_CC]
    target_tuple = tuple(sorted((name.lower() for name in target_tuple)))                                               # emp = tuple(sorted((row['First name'].lower(), row['Last name'].lower())))
    if target_tuple in list_of_tuples:
        return True
    else:
        return False
# def compare_dates(date1, date2):
#     """
#     Compares two dates, which can be in string format or datetime/date objects, ignoring the time component.
#     Returns True if the dates represent the same calendar day, False otherwise.
#     Handles None and non-string types by attempting to convert them to strings for parsing.
#     Parameters:
#     - date1: The first date, can be a string, datetime/date object, or None.
#     - date2: The second date, can be a string, datetime/date object, or None.
#     Returns:
#     - True if dates are the same day, False otherwise.
#     """
#     try:
#         # Convert non-None and non-date/datetime inputs to string for parsing
#         if date1 is not None and not isinstance(date1, (datetime, date)):
#             date1 = str(date1)
#         if date2 is not None and not isinstance(date2, (datetime, date)):
#             date2 = str(date2)
#         # Parse the dates to datetime objects if they are strings
#         if isinstance(date1, str):
#             date1 = parse(date1, dayfirst=True)
#         if isinstance(date2, str):
#             date2 = parse(date2, dayfirst=True)
#         # Return False if either date is None after handling
#         if date1 is None or date2 is None:
#             return False
#         # Compare the date components of the datetime objects
#         return date1.date() == date2.date()
#         # if not kwargs:
#         #     return date1.date() == date2.date()
#         # else:
#         #     return [date1.date(), date2.date()]
#     except Exception as e:
#         # Log or handle the exception if needed
#         # print(f"Error parsing dates: {e}")
#         return False
# def check_pairs_insensitive(mapping, column1, column2):
#     """
#     Checks if each pair formed by corresponding elements from column1 and column2
#     meets its "twin" as defined in the mapping, ignoring case sensitivity and handling None or non-string values.
#     The function now also supports mappings where the values are lists of possible matches.
#     Parameters:
#     - mapping: A dictionary where keys are the first element of the pair (in lowercase) and values are the second element or a list of second elements (also in lowercase).
#     - column1: A list of the first elements in the pairs.
#     - column2: A list of the second elements in the pairs.
#
#     Returns:
#     - A list of boolean values indicating if each pair meets its "twin", ignoring case sensitivity.
#     """
#     results = []
#     # Adjust the mapping to ensure all keys are lowercase and values are lists of lowercase strings
#     mapping_lower = {}
#     for k, v in mapping.items():
#         if k is not None:
#             key_lower = str(k).lower()
#             # Ensure every value is a list for consistency
#             value_lower = [str(item).lower() for item in v] if isinstance(v, list) else [str(v).lower()]
#             mapping_lower[key_lower] = value_lower
#     for x, y in zip(column1, column2):
#         # Ensure x and y are not None, then convert to string and lowercase for comparison
#         x_str = "" if x is None else str(x).lower()
#         y_str = "" if y is None else str(y).lower()
#         # Retrieve the list of possible matches for x_str
#         possible_matches = mapping_lower.get(x_str, [])
#         # Check if y_str is in the list of possible matches
#         match_found = y_str in possible_matches
#         results.append(match_found)
#     return results
def add_dict_column_and_export(df, columns_list):
    """
    Add a new column to the DataFrame with a dictionary of specified columns,
    and export the DataFrame to a CSV file.

    :param df: The original DataFrame.
    :param columns_list: List of column names to include in the dictionary.
    :param csv_file_name: Name of the CSV file to export.
    """
    # Function to create a dictionary from a row for the specified columns
    def row_to_dict(row):
        return {col: row[col] for col in columns_list}
    # Apply the function to each row and assign the result to a new column 'other'
    df.loc[:, 'other'] = df.apply(row_to_dict, axis=1)
    return df
def create_dict_from_df(df, column_x):
    """
    Create a dictionary from a DataFrame based on a specified column's values.

    :param df: The original DataFrame.
    :param column_x: The column name based on which to filter and create the dictionary.
    :return: A dictionary with unique values from column_x as keys and filtered DataFrames as values.
    """
    unique_values = df[column_x].unique()
    dict_from_df = {value: df[df[column_x] == value] for value in unique_values}
    return dict_from_df
def excel_file_editor(filePath, fileName):
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter
    if not filePath.strip():
        # If filePath is empty, use the current directory
        fullPath = fileName
        styledFileName = f"styled_{fileName}"
    else:
        # Construct the full path using os.path.join for better compatibility
        fullPath = os.path.join(filePath, fileName)
        styledFileName = os.path.join(filePath, f"styled_{fileName}")
    # Load the workbook and select the active worksheet
    workbook = load_workbook(filename=fullPath)
    sheet = workbook.active
    # Define the fill color and font style for the headers
    header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow fill
    header_font = Font(bold=True)  # Bold font
    # Assuming your headers are in the first row, apply styling
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        # Optional: Auto-adjust column width
        sheet.column_dimensions[get_column_letter(col)].auto_size = True
        # Note: auto_size attribute does not exist in openpyxl. You might need to set column widths manually if needed.
    workbook.save(filename=styledFileName)
    # Open the styled Excel file
    if platform.system() == "Windows":
        os.startfile(styledFileName)
    elif platform.system() == "Darwin":  # macOS
        subprocess.call(["open", styledFileName])
    else:  # Assume Linux or similar
        subprocess.call(["xdg-open", styledFileName])
    return
#   -------------------------------------------------------------------------------------------------------------      #
#   -------------------------------------------------------------------------------------------------------------      #
#   -------------------------------------------------------------------------------------------------------------      #
def hc_df_preprocess(df):
    def hc_extract_employees_df(df):
        hc_employees_df_column_names = df.iloc[3, 0:10].values.tolist()
        hc_employees_df = df.iloc[4:, 0:10].copy()
        hc_employees_df.columns = hc_employees_df_column_names
        hc_employees_df = hc_employees_df.reset_index().rename(columns={'index': 'hc_original_index'})
        # List of column for creating the identifier & Applying
        hc_employees_df.insert(0, 'db_id', hc_employees_df.apply(lambda row: create_identifier(row, ['First name', 'Last name']), axis=1))  #   db_id_columns_list = ['First name', 'Last name']
        print(f'hc_employees_df: {hc_employees_df.info()}') # hc_employees_df.to_csv("01_temp_hc_employees_df.csv", index=False)
        return hc_employees_df
    def hc_extract_periods_df(df):
        hc_periods_df_column_names = df.iloc[3, 11:52].values.tolist()
        new_date = hc_periods_df_column_names[-1] + timedelta(days=1)
        hc_periods_df_column_names.append(new_date)
        hc_formatted_column_names = [pd.to_datetime(date).strftime('%d/%m/%Y') for date in hc_periods_df_column_names]         # Convert each column name to datetime and then format it as 'DD/MM/YYYY'
        hc_periods_df = df.iloc[4:, 11:53].copy()
        hc_periods_df.columns = hc_formatted_column_names
        hc_periods_df = hc_periods_df.reset_index().rename(columns={'index': 'hc_original_index'})
        #   ---------------------------------------------------------------------------------------------------------      #
        # hc_periods_df.to_csv("02_temp_hc_periods_df.csv", index=False)
        return hc_periods_df
    def hc_periods_extractor(df):
        def hc_aggregate_dates(row):
            periods = []                                                                                                # Initialize an empty list to store the tuples of start and end dates
            start_date = None
            for date, value in row.iteritems():                                                                         ### Iterate over the row items (date and value)
                if value == 1 and start_date is None:                                                                   ### Check if we find the start of a period (value is 1)
                    start_date = date                                                                                   # Mark the start date
                elif value != 1 and start_date is not None:                                                             ### Check if we find the end of a period (value is 0 or NaN after a start date is set)
                    new_date = datetime.strptime(prev_date, '%d/%m/%Y') + timedelta(days=1)
                    new_date_str = new_date.strftime('%d/%m/%Y')
                    end_date = new_date_str
                    periods.append((start_date, end_date))                                                              # Append the period as a tuple
                    start_date = None                                                                                   # Reset the start date for the next period
                prev_date = date                                                                                        # Keep track of the previous date for end date marking
            if start_date is not None:                                                                                  ### Check if the last value was 1, indicating an ongoing period
                periods.append((start_date, prev_date))                                                                 # The period goes until the last date
            return periods
        df['hc_per_tup'] = df.apply(hc_aggregate_dates, axis=1)
        df['1st_hc'] = df['hc_per_tup'].apply(lambda x: x[0] if len(x)>0 else None)
        df['2nd_hc'] = df['hc_per_tup'].apply(lambda x: x[1] if len(x)>1 else None)
        df['3rd_hc'] = df['hc_per_tup'].apply(lambda x: x[2] if len(x)>2 else None)
        # Extracting the 'in' and 'out' for the 1st hotel stay
        df['hc_1st_hotel_in'] = df['hc_per_tup'].apply(lambda x: x[0][0] if x and len(x) > 0 else None)
        df['hc_1st_hotel_out'] = df['hc_per_tup'].apply(lambda x: x[0][1] if x and len(x) > 0 else None)
        # Extracting the 'in' and 'out' for the 2nd hotel stay
        df['hc_2nd_hotel_in'] = df['hc_per_tup'].apply(lambda x: x[1][0] if x and len(x) > 1 else None)
        df['hc_2nd_hotel_out'] = df['hc_per_tup'].apply(lambda x: x[1][1] if x and len(x) > 1 else None)
        # Extracting the 'in' and 'out' for the 3rd hotel stay
        df['hc_3rd_hotel_in'] = df['hc_per_tup'].apply(lambda x: x[2][0] if x and len(x) > 2 else None)
        df['hc_3rd_hotel_out'] = df['hc_per_tup'].apply(lambda x: x[2][1] if x and len(x) > 2 else None)
        # Change for comperhasion (not efficient!)
        df['hc_1st_hotel_in'] = pd.to_datetime(df['hc_1st_hotel_in'], format='%d/%m/%Y').dt.strftime('%Y-%m-%d')
        df['hc_1st_hotel_out'] = pd.to_datetime(df['hc_1st_hotel_out'], format='%d/%m/%Y').dt.strftime('%Y-%m-%d')
        df['hc_2nd_hotel_in'] = pd.to_datetime(df['hc_2nd_hotel_in'], format='%d/%m/%Y').dt.strftime('%Y-%m-%d')
        df['hc_2nd_hotel_out'] = pd.to_datetime(df['hc_2nd_hotel_out'], format='%d/%m/%Y').dt.strftime('%Y-%m-%d')
        df['hc_3rd_hotel_in'] = pd.to_datetime(df['hc_3rd_hotel_in'], format='%d/%m/%Y').dt.strftime('%Y-%m-%d')
        df['hc_3rd_hotel_out'] = pd.to_datetime(df['hc_3rd_hotel_out'], format='%d/%m/%Y').dt.strftime('%Y-%m-%d')
        list_to_exp = ['hc_original_index', 'hc_per_tup',
                        '1st_hc', 'hc_1st_hotel_in', 'hc_1st_hotel_out',
                        '2nd_hc', 'hc_2nd_hotel_in', 'hc_2nd_hotel_out',
                        '3rd_hc', 'hc_3rd_hotel_in', 'hc_3rd_hotel_out']
        # df[list_to_exp].to_csv("03_temp_periods.csv", index=False)                                                      # df[['hc_original_index', 'hc_per_tup', '1st_hc', '2nd_hc', '3rd_hc']].to_csv("03_temp_periods.csv", index=False)
        return df[list_to_exp]
    def export_HC_for_kevin():
        cwt_data = pd.read_excel(io=f'./rawData/emails_cwt.xlsx', sheet_name="Sheet1")
        cwt_data['Employee ID'] = "'" + cwt_data['Employee ID'].astype(str).str.replace('\D', '', regex=True)
        cwt_data['Email - Work'] = cwt_data['Email - Work'].str.lower()
        if HC_TARGET_FOLDER.exists() and HC_TARGET_FOLDER.is_dir() and not LOCAL_HC_FLAG:                               # Check if the folder exists
            os.chdir(HC_TARGET_FOLDER)                                                                                  # Change the current working directory to target folder
            files = os.listdir('.')
            if 'HC all attendees.xlsx' in files:
                hc_raw_data = pd.read_excel(io='HC all attendees.xlsx', sheet_name=HC_SHEET_NAME)                          # Getting updated HC data.
                hc_efrat_raw_data = hc_extract_employees_df(df=hc_raw_data)
                hc_efrat_raw_data['Employee direct manager email'] = hc_efrat_raw_data['Employee direct manager email'].str.lower()
                hc_efrat_raw_data['Employee Email'] = hc_efrat_raw_data['Employee Email'].str.lower()
                hc_efrat_raw_data['First name'] = hc_efrat_raw_data['First name'].str.lower()
                hc_efrat_raw_data['Last name'] = hc_efrat_raw_data['Last name'].str.lower()
                hc_efrat_raw_data['HP Employee#'] = "'" + hc_efrat_raw_data['HP Employee#'].astype(str).str.replace('\D', '', regex=True)
                temp_df = hc_efrat_raw_data.copy()
                filtered_data = hc_efrat_raw_data[(hc_efrat_raw_data['Employee Email'].notna()) & (hc_efrat_raw_data['Employee Email'].str.strip() != '') & (hc_efrat_raw_data['Employee Email'].str.strip() != 'tbd')]
                merge_filtered_data = pd.merge(filtered_data, cwt_data,
                                               left_on=['Employee Email'], right_on=['Email - Work'],
                                               how='inner', suffixes=('_df1', '_df2'))
                filtered_data_cols = ['Last name', 'First name', 'Employee Email', 'Employee ID']
                os.chdir(ORG_CWD)
                attendees_data_folder = "attendeesData"
                if not os.path.exists(f"./{attendees_data_folder}"):
                    os.makedirs(attendees_data_folder)
                f = os.path.join(attendees_data_folder, f"{today_date}_hc_efrat_raw_data.csv")
                temp_df.to_csv(f, index=False)
                f = os.path.join(attendees_data_folder, f"{today_date}_drupa_attendees.csv")
                merge_filtered_data[filtered_data_cols].to_csv(f, index=False)
                return hc_raw_data, merge_filtered_data
            else:
                print(f"The folder {HC_TARGET_FOLDER} does not exist or is not a directory.")
            os.chdir(ORG_CWD)
        return
    def employeesNumber_from_rawData(df1, df2):
        # Drop duplicates in df2 based on the lowercase email to ensure only the first match is used
        df2_unique = df2.drop_duplicates(subset=['email'], keep='first')
        # Merge df1 with the modified df2_unique, matching on the lowercase email columns
        merged_df = pd.merge(df1, df2_unique[['email', 'manEmpNum']], left_on='Employee Email', right_on='email', how='left')
        merged_df['manEmpNum'] = "'" + merged_df['manEmpNum'].astype(str).str.replace('\D', '', regex=True)
        return merged_df
    try:
        hc_raw_data,  hc_for_kevin_df = export_HC_for_kevin()
        print(hc_for_kevin_df.columns.to_list())
        print(f'####################\nhc_for_kevin_df:')
        hc_for_kevin_df.info()
        # TODO: NOTE: gathering the last update automated:
        df = hc_raw_data.copy()
    except Exception as e:
        print(f"Error0 - None update HC file - using the uploaded file.:\n{e}")
    hc_employees_df = hc_extract_employees_df(df=df)
    hc_periods_df = hc_periods_extractor(df=hc_extract_periods_df(df=df))
    hc_comb_df = pd.merge(hc_employees_df, hc_periods_df, on='hc_original_index', how='outer')
    hc_comb_df['source_HC'] = "HC"
    hc_comb_df['Employee direct manager email'] = hc_comb_df['Employee direct manager email'].str.lower()
    hc_comb_df['Employee Email'] = hc_comb_df['Employee Email'].str.lower()
    hc_comb_df['First name'] = hc_comb_df['First name'].str.lower()
    hc_comb_df['Last name'] = hc_comb_df['Last name'].str.lower()
    hc_comb_df['HP Employee#'] = "'" + hc_comb_df['HP Employee#'].astype(str).str.replace('\D', '', regex=True)
    # TODO: read the rawData_with the empNum - drop from flight df the emp.Num that not relevant to the empNum from the HC! then continue regular
    file_path = './rawData/employees_num_emails.xlsx'
    sheet_name = 'emp'  # Cols: 'manEmpNum', 'email', 'HC_index'
    dtype_dict = {'manEmpNum': str}
    employees_num_emails_df = pd.read_excel(io=file_path, sheet_name=sheet_name, dtype=dtype_dict)
    employees_num_emails_df['email'] = employees_num_emails_df['email'].str.lower()
    hc_comb_df = employeesNumber_from_rawData(df1=hc_comb_df, df2=employees_num_emails_df)
    hc_comb_df.rename(columns={'manEmpNum': 'EmpNum'}, inplace=True)
    COLS_FOR_LIMOR.insert(0, 'EmpNum')
    hc_comb_df.to_csv('00_employeesNumber_from_rawData.csv', index=False)
    # TODO: Done
    os.chdir(ORG_CWD)
    print(f'####################\nhc_comb_df:')
    hc_comb_df.info()
    print(hc_comb_df.columns.to_list())
    # hc_comb_df.to_csv("04_temp_hc_comb_df.csv", index=False)                                                          # hc_comb_df.to_csv("05_temp_hc_comb_df.csv", index=False)
    return hc_comb_df
#   -------------------------------------------------------------------------------------------------------------      #
def cnpi_df_preprocess(df):
    raw_data = df.copy()
    # raw_data['1st_cnpi'] = list(zip(pd.to_datetime(raw_data['1st Hotel Check in']).dt.strftime('%d/%m/%Y'), pd.to_datetime(raw_data['1st Hotel Check out']).dt.strftime('%d/%m/%Y')))
    raw_data['1st_cnpi'] = [
        (pd.to_datetime(start).strftime('%d/%m/%Y'), pd.to_datetime(end).strftime('%d/%m/%Y'))
        if pd.notnull(start) and pd.notnull(end) else None
        for start, end in zip(raw_data['1st Hotel Check in'], raw_data['1st Hotel Check out'])
        ]
    # raw_data['2nd_cnpi'] = list(zip(pd.to_datetime(raw_data['2nd Hotel Check in']).dt.strftime('%d/%m/%Y'), pd.to_datetime(raw_data['2nd Hotel Check out']).dt.strftime('%d/%m/%Y')))
    raw_data['2nd_cnpi'] = [
        (pd.to_datetime(start).strftime('%d/%m/%Y'), pd.to_datetime(end).strftime('%d/%m/%Y'))
        if pd.notnull(start) and pd.notnull(end) else None
        for start, end in zip(raw_data['2nd Hotel Check in'], raw_data['2nd Hotel Check out'])
        ]
    # raw_data['3rd_cnpi'] = list(zip(pd.to_datetime(raw_data['3rd Hotel Check in']).dt.strftime('%d/%m/%Y'), pd.to_datetime(raw_data['3rd Hotel Check out']).dt.strftime('%d/%m/%Y')))
    raw_data['3rd_cnpi'] = [
        (pd.to_datetime(start).strftime('%d/%m/%Y'), pd.to_datetime(end).strftime('%d/%m/%Y'))
        if pd.notnull(start) and pd.notnull(end) else None
        for start, end in zip(raw_data['3rd Hotel Check in'], raw_data['3rd Hotel Check out'])
        ]
    raw_data['cnpi_per_tup'] = raw_data.apply(lambda row: [row['1st_cnpi'], row['2nd_cnpi'], row['3rd_cnpi']], axis=1)
    #   ---------------------------------------------------------------------------------------------------------      #
    # List of column for creating the identifier & Applying
    raw_data['db_id'] = raw_data.apply(lambda row: create_identifier(row, ['firstName', 'lastName']), axis=1)           # db_id_columns_list = ['firstName', 'lastName']
    #   ---------------------------------------------------------------------------------------------------------      #
    raw_data['firstName'] = raw_data['firstName'].str.lower()
    raw_data['lastName'] = raw_data['lastName'].str.lower()
    raw_data['email'] = raw_data['email'].str.lower()
    canapi_df = raw_data.copy()
    canapi_df['source_CP'] = "CP"
    canapi_df_columns_to_exp = ['db_id', 'source_CP', 'firstName', 'lastName', 'email', 'type', 'ticketName', 'Role at Show', 'Country', 'cnpi_per_tup', '1st_cnpi', '2nd_cnpi', '3rd_cnpi',
                                'How will you be travelling to Germany?', 'jobTitle', 'company', 'Organization', 'phone_number',
                                 'Passport Number', 'Passport Expiry Date', 'gender', 'HP Job Title',
                                 'Tshirt Size', 'Manager Name', 'Manager Email Address', 'Type', 'Company Name', 'Job Title',
                                 '1st Hotel Check in', '1st Hotel Check out',
                                 '2nd Hotel Check in', '2nd Hotel Check out',
                                 '3rd Hotel Check in', '3rd Hotel Check out',
                                 'Flight 1', 'Flight 2', 'Flight 3',
                                 'Add additional flight? [Flight 2]', 'Add additional flight? [Flight 3]',
                                 'Flight 1: Arrival Airport', 'Flight 1: Arrival Date', 'Flight 1: Arrival Time', 'Flight 1: Flight Number',
                                 'Flight 2: Arrival Airport', 'Flight 2: Arrival Date', 'Flight 2: Arrival Time', 'Flight 2: Flight Number',
                                 'Flight 3: Arrival Airport', 'Flight 3: Arrival Date', 'Flight 3: Arrival Time', 'Flight 3: Flight Number']
    # existing_columns = [col for col in canapi_df_columns_to_exp if col in raw_data.columns]
    # raw_data[existing_columns].to_csv("05_temp_canapi_df.csv", index=False)
    print(f'####################\ncanapi_df:')
    canapi_df.info()
    return canapi_df
#   -------------------------------------------------------------------------------------------------------------      #
#   -------------------------------------------------------------------------------------------------------------      #
#   -------------------------------------------------------------------------------------------------------------      #
class CustomButton:
    def __init__(self, parent, text, description, command, row, column, identifier, width=20, height=2, padx=10, pady=10):
        """
        Sticky:
        "n": aligns to top of the cell.
        "s": aligns to bottom of the cell.
        "e": aligns to right of the cell.
        "w": aligns to left of the cell.
        """
        self.parent = parent  # Save parent to use for scheduling GUI operations
        self.button = tk.Button(parent, text=text, command=lambda: self.execute_in_thread(command, identifier), width=width, height=height)
        self.button.grid(row=row, column=column, padx=padx, pady=pady, sticky="new")
        # Button Label definition
        self.info_label = tk.Label(parent, text=description)
        self.info_label.grid(row=row, column=column+1, padx=padx, pady=pady, sticky="nw")
        # Initialize an empty DataFrame
        self.data = pd.DataFrame()
    def execute_in_thread(self, command, identifier):
        """Executes the given command in a separate thread."""
        Thread(target=lambda: command(identifier, self.post_to_gui_thread), daemon=False).start()
    def post_to_gui_thread(self, func, *args, **kwargs):
        """Post a task to the Tkinter main thread."""
        self.parent.after(0, lambda: func(*args, **kwargs))
#   -------------------------------------------------------------------------------------------------------------      #
class ExcelFileSelector:
    def __init__(self, tab):
        self.tab = tab                                                                                                  # Use this self.tab instead of self.root # self.tab.title("DRUPA App") - the tab's title instead
        self.buttons = {}
        self.fin_df = None
        self.raw_data = None
        self.last_selected_button = None
        self.hc_comb_df = None
        self.canapi_df = None
        self.cwt_df = None
        self.ofir_df = None
        self.initialize_ui()
        self.auto_adjust_grid()
    def initialize_ui(self):
        # HC Button
        self.buttons["HC"] = CustomButton(parent=self.tab, text="Select Head Count File", description="Load last version of HC File",
                                          command=self.data_files_loader, identifier="HC",                              row=0, column=0)
        # CANNAPI Button
        self.buttons["CANAPII"] = CustomButton(parent=self.tab, text="Select Canapii File", description="Load last version from Canapii",
                                               command=self.data_files_loader, identifier="CANAPII",                    row=1, column=0)
        # Registration Button
        self.buttons["REG"] = CustomButton(parent=self.tab, text="Registration Analysis", description="Analysis and export",
                                           command=self.registration_analysis, identifier="REG",                        row=2, column=0)

        self.buttons["LOADIND_SPLT"] = CustomButton(parent=self.tab, text="Select SPLT File", description="Load data for SPLT",
                                                    command=self.data_files_loader, identifier="LOADIND_SPLT",          row=3, column=0)
        # Splitting column
        self.buttons["SPLT"] = CustomButton(parent=self.tab, text="Splitting", description="Splitting traveler",
                                            command=self.splitting_traveler_number_name, identifier="SPLT",             row=4, column=0)
        # Loading CWT
        self.buttons["CWT"] = CustomButton(parent=self.tab, text="CWT data", description="CWT Data",
                                            command=self.data_files_loader, identifier="CWT",                           row=5, column=0)
        # Loading OFIR
        self.buttons["OFIR"] = CustomButton(parent=self.tab, text="OFIR data", description="OFIR Data",
                                            command=self.data_files_loader, identifier="OFIR",                          row=6, column=0)
        # Loading FLIGHT
        self.buttons["FLIGHT"] = CustomButton(parent=self.tab, text="FLIGHT analysis", description="Flight Analysis",
                                            command=self.flights_analysis, identifier="FLIGHT",                        row=7, column=0)
        # Loading Employees
        self.buttons["Employees"] = CustomButton(parent=self.tab, text="Employees extractor", description="Employees extractor",
                                            command=self.employees_extractor, identifier="Employees",                   row=8, column=0)
    #   TODO: ---------------------------------------------------------------------------------------------------      #
    #   TODO: ---------------------------------------------------------------------------------------------------      #
    #   TODO: ---------------------------------------------------------------------------------------------------      #
    def auto_adjust_grid(self):
        max_row, max_column = 0, 0
        for widget in self.tab.winfo_children():
            grid_info = widget.grid_info()
            max_row = max(max_row, grid_info['row'])
            max_column = max(max_column, grid_info['column'])
        for row in range(max_row + 1):
            self.tab.grid_rowconfigure(row, weight=1)
        for column in range(max_column + 1):
            self.tab.grid_columnconfigure(column, weight=1)
    def update_gui_after_processing(self, **kwargs):                                                                    # Safely update the GUI from the main thread
        """self.tab.after(0, lambda: self.update_gui_after_processing(message="Analysis done"))"""
        try:
            if self.buttons[self.last_selected_button].data is not None:                                                # selected_button = self.buttons.get(self.last_selected_button)
                num_records = len(self.buttons[self.last_selected_button].data)
                if self.last_selected_button=="HC":
                    num_records-=10
                message = kwargs.get('message', f"Loaded {num_records} records.")
                self.buttons[self.last_selected_button].info_label.config(text=message)                                 # self.buttons[self.last_selected_button].info_label.config(text=f"Loaded {num_records} records.")        # selected_button.info_label.config(text=f"Loaded {len(self.raw_data)} records.")
        except Exception as e:
            error_message = f"Updating UI failed due to:\n{e}"
            print(error_message)
            self.tab.after(0, lambda: messagebox.showerror(f"Error - {self.last_selected_button}", error_message))
    def data_files_loader(self, identifier, post_to_gui):
        self.last_selected_button = identifier
        try:
            file_path = tk.filedialog.askopenfilename(filetypes=[("Excel files", "*")])
            if not file_path:                                                                                           # User cancelled the dialog
                return
            else:
                if identifier == "HC":
                    raw_data = pd.read_excel(io=file_path, sheet_name=HC_SHEET_NAME)
                    self.hc_comb_df = hc_df_preprocess(df=raw_data)                                                     # hc_employees_df = hc_extract_employees_df(df=raw_data)    ##########    # hc_periods_df = hc_periods_extractor(df=hc_extract_periods_df(df=raw_data))     ##########                     # hc_comb_df = pd.merge(hc_employees_df, hc_periods_df, on='hc_original_index', how='outer').to_csv("05_temp_hc_comb_df.csv", index=False)   # hc_comb_df.to_csv("05_temp_hc_comb_df.csv", index=False)
                    print("self.hc_comb_df")
                    print(self.hc_comb_df.info())
                elif identifier == "CANAPII":
                    raw_data = pd.read_csv(filepath_or_buffer=file_path)
                    self.canapi_df = cnpi_df_preprocess(df=raw_data)
                    print("self.canapi_df")
                    print(self.canapi_df.info())
                elif identifier == "LOADIND_SPLT":
                    raw_data = pd.read_excel(io=file_path)
                    self.for_splt_df = raw_data.copy()
                elif identifier == "CWT":
                    df_temp = pd.read_excel(file_path)
                    # Step 2: Find the row with 'Traveler Name' in the first column to identify the header
                    header_row_index = df_temp[df_temp.iloc[:, 11] == 'Transactions'].index.min() + 1
                    # Step 3: Find the row with 'Grand Total' to identify the last useful row (exclude this row)
                    last_row_index = df_temp[df_temp.iloc[:, 11].notna()].index.max()
                    # Now, read the file again with the identified header row and use nrows to limit the read
                    raw_data = pd.read_excel(file_path, header=header_row_index, nrows=last_row_index - header_row_index)
                    self.cwt_df = raw_data.copy()
                    print("self.cwt_df")
                    print(self.cwt_df.info())
                    pass
                elif identifier == "OFIR":
                    raw_data = pd.read_excel(io=file_path)
                    self.ofir_df = raw_data.copy()
                    pass
                #   ------------------------------------------------------------------------------------------------   #
                self.buttons[identifier].data = raw_data.copy()                                                         # Assign the raw_data for each file for further usage
                self.tab.after(0, lambda: self.update_gui_after_processing())
        except Exception as e:
            error_message = f"Failed to open filed due to:\n{e}"
            print(error_message)
            self.tab.after(0, lambda: messagebox.showerror(f"Error1 - {identifier}", error_message))                     # self.root.after(0, lambda: messagebox.showerror("Error", f"Failed to open file {e}"))
    #   TODO: ---------------------------------------------------------------------------------------------------      #
    #   TODO: ---------------------------------------------------------------------------------------------------      #
    #   TODO: ---------------------------------------------------------------------------------------------------      #
    def splitting_traveler_number_name(self, identifier, post_to_gui):
        self.last_selected_button = identifier
        df = self.for_splt_df.copy()
        # Extract the column values
        passenger_column = df['שם נוסע']
        # Split the values
        traveler_ids = passenger_column.str.extract(r'(\d+)').squeeze()
        traveler_names = passenger_column.str.replace(r'(\d+\s+)', '', regex=True).squeeze()
        # Add new columns to the DataFrame
        df['empNumber'] = traveler_ids
        df['empNumber'] = "'" + df['empNumber'].astype(str).str.replace('\D', '', regex=True)
        df['empName'] = traveler_names.str.lower()
        df[['lastName', 'firstName']] = df['empName'].str.split('/', expand=True)
        hc_cwt_data = pd.read_excel(io=f'./rawData/emails_cwt.xlsx')
        hc_cwt_data['Employee ID'] = "'" + hc_cwt_data['Employee ID'].astype(str).str.replace('\D', '', regex=True)
        merged_df = pd.merge(df, hc_cwt_data, left_on='empNumber', right_on='Employee ID', how='inner')
        # all_cols = [
        #                         'מספר דוקט', 'PNR', 'שם נוסע', 'חשבונית', 'תאריך יציאה', 'תאריך חזרה', 'ספק מזמין', 'יעד', 'IL', 'ימים מראש',
        #                         'סה"כ ימי טיסה', 'ברוטו + מיסים - הנחה', 'TOTAL', 'Cost Center', 'empNumber', 'empName', 'lastName',
        #                          'firstName', 'Employee', 'Employee ID', 'Legal Name', 'Legal Name - Last Name', 'Legal Name - First Name',
        #                      'Legal Name in Reporting Display Format', 'Preferred Name', 'Preferred Name - Last Name',
        #                      'Preferred Name - First Name', 'Preferred Name in Reporting Display Format', 'Email - Work',
        #                      'Job Family Group', 'Compensation Grade Profile', 'Business lvl 1 (Group)', 'Business lvl 2 (Unit)',
        #                      'Business lvl 3 (Org Chart)', 'Business lvl 4 (MRU)', 'Business lvl 5 (Sub-MRU)', 'Cost Centers', 'S4 CC',
        #                      'QAD CC', 'SUB ACCOUNT', 'Manager ID', 'Manager Name', 'Manager Email', 'Management Chain - Level 01',
        #                      'Management Chain - Level 02', 'Management Chain - Level 03', 'Management Chain - Level 04',
        #                      'Management Chain - Level 05', 'Management Chain - Level 06', 'Management Chain - Level 07', 'Reporting Level',
        #                      'Theater Code', 'Theater Description', 'Country Code', 'Country Description', 'Location',
        #                      'Location Address - City', 'Location Address - State/Province', 'Location Address - Postal Code',
        #                      'Location Address - Country/Region', 'Phone - Primary Work', 'Workplace Type', 'Workplace', 'Acquisition Code',
        #                      'Acquisition Desc', 'Union Membership Type', 'Employee Type', 'Time Type', 'Employment Status', 'Active (Y/N)',
        #                      'Terminated (Y/N)', 'Leave (Y/N)', 'Continuous Service Date', 'Length of Service (in years)',
        #                      'Original Hire Date', 'Hire/Re-Hire Date', 'Primary Home Address', 'Primary Home Address - City',
        #                      'Primary Home Address - State/Province', 'Primary Home Address - Postal Code',
        #                      'Primary Home Address - Country', 'Phone - Primary Home', 'Report Effective Date and Time']
        cols_to_include = [     'מספר דוקט', 'PNR', 'שם נוסע', 'חשבונית', 'תאריך יציאה', 'תאריך חזרה', 'ספק מזמין', 'יעד', 'IL', 'ימים מראש',
                                'סה"כ ימי טיסה', 'ברוטו + מיסים - הנחה', 'TOTAL', 'Cost Center', 'empNumber', 'empName', 'lastName',
                                 'firstName','Email - Work','Manager ID', 'Manager Name', 'Manager Email']
        merged_df[cols_to_include].to_csv("split_merged_column_df.csv", index=False, encoding='utf-8-sig')
        print(f'####################\nsplit_merged_column_df:')
        merged_df[cols_to_include].info()
    #   TODO: ---------------------------------------------------------------------------------------------------      #
    #   TODO: ---------------------------------------------------------------------------------------------------      #
    #   TODO: ---------------------------------------------------------------------------------------------------      #
    def registration_analysis(self, identifier, post_to_gui):
        self.last_selected_button = identifier
        def non_fully_defined_hc_ppl_func(df):
            crucial_column_to_be_defined = ['First name', 'Last name', 'Employee Email', 'Workstream', 'HP Employee#']
            non_defined_hc_df = df[df[crucial_column_to_be_defined].apply(lambda x: x.str.lower().isin(['tbd']) | x.isin(['', np.nan])).any(axis=1)].copy()
            return non_defined_hc_df
        def merging_df_func(hc_df, cnpi_df):
            hc_df.to_csv("00_hc_df.csv", index=True)
            print(f'####################\t00_hc_df:')
            hc_df.info()
            cnpi_df.to_csv("00_cnpi_df.csv", index=True)
            print(f'####################\t00_cnpi_df:')
            cnpi_df.info()
            subset_to_drop_as_nulls_hc = ["Last name", "First name", "Employee Email", "Workstream"]                    # subset_to_drop_as_nulls_hc = ["Last name", "First name", "Employee Email", "Employee direct manager email", "Workstream"]
            hc_df_cleaned = hc_df.dropna(subset=subset_to_drop_as_nulls_hc, how='all')
            columns_to_check_TBDs = ['First name', 'Last name', 'Employee Email']                                       # columns_to_check_TBDs = ['First name', 'Last name', 'firstName', 'lastName', 'Employee Email', 'email']
            hc_df_cleaned = hc_df_cleaned[~hc_df_cleaned[columns_to_check_TBDs].apply(lambda x: x.str.lower().isin(['tbd']).any(), axis=1)].copy()  # analyzed_reg_df = analyzed_reg_df[~analyzed_reg_df[columns_to_check_TBDs].apply(lambda x: x.str.lower().isin(['tbd']).any(), axis=1)].copy()
            hc_df_cleaned.to_csv("01_hc_df_cleaned.csv", index=True)
            print(f'####################\t01_hc_df_cleaned:')
            hc_df_cleaned.info()
            # TODO: Outer data #  -----------------------------------------------------------------------------------  #
            CH_CP_MERGED_BY_EMAIL_OUTER = pd.merge(hc_df_cleaned, cnpi_df, left_on=['Employee Email'], right_on=['email'], how='outer', suffixes=('_df1', '_df2'))
            # CH_CP_MERGED_BY_EMAIL_OUTER.to_csv("temp_CH_CP_MERGED_BY_EMAIL_OUTER.csv", index=True)
            # print(f'####################\nCH_CP_MERGED_BY_EMAIL_OUTER:')
            # CH_CP_MERGED_BY_EMAIL_OUTER.info()
            # TODO: Two levels of merging #  ------------------------------------------------------------------------  #
            inner_merge_email = pd.merge(hc_df_cleaned, cnpi_df, left_on=['Employee Email'], right_on=['email'], how='inner', suffixes=('_df1', '_df2'))
            # inner_merge_email.to_csv("temp_CH_CP_MERGED_BY_EMAIL_INNER.csv", index=False)
            # print(f'####################\ninner_merge_email:')
            # inner_merge_email.info()
                # TODO:  1. Identify IDs that have not been matched by email   #  ----------------------------------   #
            unmatched_ids = hc_df_cleaned.loc[~hc_df_cleaned['db_id'].isin(inner_merge_email['db_id_df1']), 'db_id']
                # TODO:  2. Filter rows in df1 and df2 for the second merge attempt    #  --------------------------   #
            df1_unmatched = hc_df_cleaned[hc_df_cleaned['db_id'].isin(unmatched_ids)]
            df2_unmatched = cnpi_df[cnpi_df['db_id'].isin(unmatched_ids)]
                # TODO:  3. Merge the unmatched rows on id #  ------------------------------------------------------   #
            merge_id = pd.merge(df1_unmatched, df2_unmatched, on='db_id', how='outer', suffixes=('_df1', '_df2'))
                # TODO:  4. Combine the results of both merges #  --------------------------------------------------   #
            final_merged_df = pd.concat([inner_merge_email, merge_id]).reset_index(drop=True)
            return final_merged_df
        def combine_hc_cnpi(row, columns):
            concatenated = f"Head Count: {row[columns[0]]}\nCanapii Reg: {row[columns[1]]}"                             # concatenated = '\n'.join([(str(row[column]).lower() if row[column] is not None else 'none') for column in columns])
            return concatenated
        def register_or_not(row, columns):
            if pd.notnull(row[columns[1]]):
                return True
            elif pd.notnull(row[columns[0]]) and not pd.notnull(row[columns[1]]):
                return False
        def exceptions_hotels_check_in_out_true_false(row, col, list_of_trues):
            if row[col] in list_of_trues:
                return True
            else:
                return ""
        def travels_checks(df):
            def custom_row_operation_level_01(row, x_col, y_col):
                if pd.isnull(row[x_col]) and pd.isnull(row[y_col]):
                    return ""
                elif pd.isnull(row[x_col]) and not pd.isnull(row[y_col]):
                    return False
                else:
                    return row[x_col] == row[y_col]
            def custom_row_operation_level_01_new(row, x_col):
                Y_COLS = [
                    '1st Hotel Check in', '1st Hotel Check out',
                    '2nd Hotel Check in', '2nd Hotel Check out',
                    '3rd Hotel Check in', '3rd Hotel Check out'
                ]
                y_cols = Y_COLS
                """
                Compares the value in x_col against values in a list of columns (y_cols).
                Returns True if x_col value matches any of the values in the y_cols, else False.
                If x_col is null, the behavior is defined by the specific condition checks.

                Parameters:
                - row: The current row of the DataFrame.
                - x_col: The name of the column to compare against others.
                - y_cols: A list of column names to compare the x_col value against.

                Returns:
                - True, False, or "" based on the comparison logic.
                """
                # If x_col is null, check against y_cols logic
                if pd.isnull(row[x_col]):
                    if all(pd.isnull(row[col]) for col in y_cols):
                        return ""  # All compared values are null
                    return False  # x_col is null but not all y_cols are null

                # Compare x_col value against all y_cols values
                for y_col in y_cols:
                    if row[x_col] == row[y_col]:
                        return True
                return False
            def custom_row_operation_level_01_new_new(row, x_col):
                New_Y_COLS = [
                    'hc_1st_hotel_in', 'hc_1st_hotel_out',
                    'hc_2nd_hotel_in', 'hc_2nd_hotel_out',
                    'hc_3rd_hotel_in', 'hc_3rd_hotel_out',
                ]
                y_cols = New_Y_COLS
                """
                Compares the value in x_col against values in a list of columns (y_cols).
                Returns True if x_col value matches any of the values in the y_cols, else False.
                If x_col is null, the behavior is defined by the specific condition checks.

                Parameters:
                - row: The current row of the DataFrame.
                - x_col: The name of the column to compare against others.
                - y_cols: A list of column names to compare the x_col value against.

                Returns:
                - True, False, or "" based on the comparison logic.
                """
                # If x_col is null, check against y_cols logic
                if pd.isnull(row[x_col]):
                    if all(pd.isnull(row[col]) for col in y_cols):
                        return ""  # All compared values are null
                    return False  # x_col is null but not all y_cols are null

                # Compare x_col value against all y_cols values
                for y_col in y_cols:
                    if row[x_col] == row[y_col]:
                        return True
                return False
            # def custom_row_operation_level_02(row, columns):
            #     if any(row[col] == False for col in columns):
            #         return False
            #     if all(row[col] == False or row[col] == "" for col in columns):
            #         return False
            #     elif all(row[col] == True or row[col] == "" for col in columns):
            #         return True
            #     return None  # Adjust this return value as needed based on your specific requirements
            def custom_row_operation_level_02(row, columns):
                if row['hotelsExceptions']:
                    return True
                if any(row[col] == False for col in columns):
                    return False
                if all(row[col] == False or row[col] == "" for col in columns):
                    return False
                elif all(row[col] == True or row[col] == "" for col in columns):
                    return True
                return None  # Adjust this return value as needed based on your specific requirements

            column_pairs = [
                ('hc_1st_hotel_in'  , '1st Hotel Check in'  , '1st_in'  ),
                ('hc_1st_hotel_out' , '1st Hotel Check out' , '1st_out' ),
                ('hc_2nd_hotel_in'  , '2nd Hotel Check in'  , '2nd_in'  ),
                ('hc_2nd_hotel_out' , '2nd Hotel Check out' , '2nd_out' ),
                ('hc_3rd_hotel_in'  , '3rd Hotel Check in'  , '3rd_in'  ),
                ('hc_3rd_hotel_out' , '3rd Hotel Check out' , '3rd_out' )
            ]
            for x_col, y_col, result_col in column_pairs:                                                               # Iterate through each specified pair and compare
                df[result_col] = df.apply(custom_row_operation_level_01, axis=1, args=(x_col, y_col))                   # df[result_col] = df[x_col] == df[y_col]
                # df[result_col] = df.apply(custom_row_operation_level_01_new, axis=1, args=(x_col,))
                # df[result_col] = df.apply(custom_row_operation_level_01_new_new, axis=1, args=(y_col,))
            dynamic_columns_hotels = tuple(pair[2] for pair in column_pairs)                                            # third_variables = [pair[2] for pair in column_pairs]
            df['Hotels_inout_approval'] = df.apply(custom_row_operation_level_02, axis=1, args=(dynamic_columns_hotels,))

            # Overall Approval
            dynamic_columns_overall = tuple(x for x in ['Registered_or_not', 'Hotels_inout_approval', 'role_flag'])
            df['overall_approval'] = df.apply(custom_row_operation_level_02, axis=1, args=(dynamic_columns_overall,))
            return df
        def cost_center_alloc(row):
            emp = (row['First name'].lower(), row['Last name'].lower())
            if not row["Workstream"] in DRUPA_CostCenter_Allocation_Workstreams:
                return "No"
            elif row["Workstream"] == "CoreTeam":
                if searching_tuple_and_list_of_tuples(target_tuple=emp, list_of_tuples=CoreTeam_emp_onDrupa_CC):
                    return "Yes"
                else:
                    return "No"
            else:
                return "Yes"
        def tshirt_sizes_validation(row):
            if row["Workstream"] in tshirts_relevancy_workstreams and pd.notnull(row['Tshirt Size']) and row['Tshirt Size'].strip() :
                return True
            else:
                return False
        def export_discrepancies_per_workstream(fin_df):
            fin_df = add_dict_column_and_export(df=fin_df, columns_list=fin_df.columns.to_list())
            cols = [
                'Employee Email', 'First name', 'Last name',  'overall_approval', 'Registered_or_not', 'role_flag', 'Workstream', 'Role at Show',
                'Hotels_inout_approval', 'hc_then_cnpi', '1st_in', '1st_out', '2nd_in', '2nd_out', '3rd_in', '3rd_out',
                'type', 'ticketName', 'gender', 'Country', 'Passport Expiry Date', 'Tshirt Size',
                'Indicate below if you have any special accommodation requests related to your religion or accessibility requirements. Please note that your request is not guaranteed and will be subject for approval.',
                'hc_per_tup', 'cnpi_per_tup', 'hc_original_index', 'other']
            fin_df[cols].to_csv(f"07_fin_df_{today_date}.csv", index=False)
            print(f'####################\n07_fin_df_:')
            fin_df[cols].info()
            #   ----------------------------------------------------------------------------------------------------   #
            workstream_fin_dfs = fin_df[cols].copy()
            result_dict = create_dict_from_df(workstream_fin_dfs, 'Workstream')                                         # Create the dictionary
            with pd.ExcelWriter(f'08_fin_df_{today_date}.xlsx', engine='xlsxwriter') as writer:                           # Create a Pandas Excel writer using XlsxWriter as the engine, outside the loop
                for key, value in result_dict.items():
                    valid_sheet_name = key[:31].replace('/', '').replace('\\', '').replace('?', '').replace('*', '').replace('[', '').replace(']', '').replace(':', '') # Ensure the sheet name is valid (Excel sheet names have a max length of 31 characters) # Also, replace characters not allowed in sheet names
                    value.to_excel(writer, sheet_name=valid_sheet_name, index=False)                                    # Write each DataFrame to a different worksheet
            os.chdir(LIMOR_FOLDER)  # Change the current working directory to target folder
            workstream_dfs_folder = "workstream_dfs"
            if not os.path.exists(f"./{workstream_dfs_folder}"):
                os.makedirs(workstream_dfs_folder)
            for key, value in result_dict.items():
                with pd.ExcelWriter(f'{workstream_dfs_folder}/{key}.xlsx', engine='xlsxwriter') as writer:              # Create a Pandas Excel writer using XlsxWriter as the engine, outside the loop
                    valid_sheet_name = key[:31].replace('/', '').replace('\\', '').replace('?', '').replace('*', '').replace('[', '').replace(']', '').replace(':', '') # Ensure the sheet name is valid (Excel sheet names have a max length of 31 characters) # Also, replace characters not allowed in sheet names
                    value.to_excel(writer, sheet_name=valid_sheet_name, index=False)
                    filtered_df_not_registered = value[value['Registered_or_not'] == False]
                    filtered_df_registered = value[value['Registered_or_not'] == True]
                    filtered_df_Discrepancies = filtered_df_registered[filtered_df_registered['overall_approval'] == False]
                    filtered_df_Discrepancies.to_excel(writer, sheet_name="Discrepancies", index=False)
                    filtered_df_not_registered.to_excel(writer, sheet_name="Not Registered", index=False)
            os.chdir(ORG_CWD)
            #   ----------------------------------------------------------------------------------------------------   #
        try:
            # 02_non_fully_defined_hc_ppl_df    --------------------------------------------------------------------   #
            non_fully_defined_hc_ppl_df = non_fully_defined_hc_ppl_func(df=self.hc_comb_df.copy())
            cols = [
                    'hc_original_index', 'HP Employee#', 'Last name', 'First name', 'Employee Email',
                    'Workstream', 'Drupa role (booth duty) / Visitor / Channel partner',
                    'Manager name', 'Employee direct manager email',
                    ]
            non_fully_defined_hc_ppl_df[cols].to_csv("02_non_fully_defined_hc_ppl_df.csv", index=False)
            print(f'####################\n02_non_fully_defined_hc_ppl_df:')
            non_fully_defined_hc_ppl_df.info()
            # 03_final_merged_df    --------------------------------------------------------------------------------   #
            analyzed_reg_df = merging_df_func(hc_df=self.hc_comb_df.copy(), cnpi_df=self.canapi_df.copy())
            analyzed_reg_df.insert(0, 'hc_then_cnpi', analyzed_reg_df.apply(lambda row: combine_hc_cnpi(row, ['hc_per_tup', 'cnpi_per_tup']), axis=1))
            analyzed_reg_df.insert(0, 'Registered_or_not', analyzed_reg_df.apply(lambda row: register_or_not(row, ['source_HC', 'source_CP']), axis=1))
            analyzed_reg_df.insert(0, 'role_flag', analyzed_reg_df.apply(lambda row: row['Role at Show'] in Workstream_Pairs_Mapping.get(row['Workstream'], []), axis=1))
            analyzed_reg_df.insert(0, 'hotelsExceptions', analyzed_reg_df.apply(lambda row: exceptions_hotels_check_in_out_true_false(row, col='Employee Email', list_of_trues=EXCLUDING_FROM_HOTELS_AND_FLIGHTS_EMAILS), axis=1))
            cols = [
                'role_flag', 'Registered_or_not', 'hc_then_cnpi', 'db_id_df1', 'hc_original_index', 'Workstream',
                'Drupa role (booth duty) / Visitor / Channel partner', 'Manager name', 'Last name', 'First name',
                'Employee Email', 'Employee direct manager email', 'HP Employee#', 'Arriving from \nAirport/Country',
                'Comments', 'hc_per_tup', '1st_hc', 'hc_1st_hotel_in', 'hc_1st_hotel_out', '2nd_hc', 'hc_2nd_hotel_in',
                'hc_2nd_hotel_out', '3rd_hc', 'hc_3rd_hotel_in', 'hc_3rd_hotel_out', 'hotelsExceptions', 'source_HC',
                'firstName', 'lastName', 'jobTitle', 'company', 'country', 'Organization', 'email', 'phone_number', 'type', 'ticketName',
                'status',
                'OtherSocialMediaURL', 'linkedInProfile', 'telNum', 'twitter', 'hotel', 'hotelCheckIn', 'hotelCheckOut',
                'Documents', 'marketingUpdates', 'First seen', 'Login time', 'Last seen', 'Registered', 'checkin',
                'Last updated date', 'Last updated by', 'Last updated attributes', 'nickName', 'gender', 'dietary',
                'Total watch time', 'Total sessions booked', '_views', 'Region', 'Country',
                'APJ - Sub Region & HP Office',
                'EMEA - Sub Region & HP Office', 'AMS - Sub Region & HP Office',
                'Languages (You may select more than one)', 'If other, please specify language', 'Are you HP employee?',
                'HP Business', 'HP Business - Industrial', 'Department/Function', 'Country Coverage',
                'Business Segment',
                'HP Job Title', 'Role at Show', 'Tshirt Size', 'Manager Name', 'Manager Email Address',
                'I am an expert in', 'Type', 'Company Name', 'Job Title',
                'Will you be attending the sales training day on May 27th?', 'Terms and Conditions',
                'What’s the Name of Your HP Sales Account Manager?', 'Additional Info', 'Nationality',
                'Passport Number',
                'Passport Expiry Date', 'Passport Issue Date', 'Passport Issued by', 'Date of Birth',
                'Dietary Restrictions', 'Life-threatening Allergy', 'AMS Taste and Talks Event Details',
                'Do you need any assistance with regards to accessibility?', 'Emergency Contact Name',
                'Emergency Contact Number', 'Visa Requirement', 'Do you need a visa for your trip?',
                'Accommodation Reminder', 'Do you require accommodation?', 'Reminder',
                'Indicate below if you have any special accommodation requests related to your religion or accessibility requirements. Please note that your request is not guaranteed and will be subject for approval.',
                '1st Hotel Check in', '1st Hotel Check out', '2nd Hotel Check in', '2nd Hotel Check out',
                '3rd Hotel Check in', '3rd Hotel Check out', 'Do you observe Shabbat?', 'Plan your Trip',
                'How will you be travelling to Germany?', 'Flight 1', 'Flight 2', 'Flight 3',
                'Add additional flight? [Flight 2]', 'Add additional flight? [Flight 3]', 'Sub Region - APJ',
                'Sub Region - AMS', 'Sub Region - EMEA', 'Your HP Office Address - EMEA', 'Flight 1: Arrival Airport',
                'Flight 1: Arrival Date', 'Flight 1: Arrival Time', 'Flight 1: Flight Number',
                'Flight 2: Arrival Airport',
                'Flight 2: Arrival Date', 'Flight 2: Arrival Time', 'Flight 2: Flight Number',
                'Flight 3: Arrival Airport',
                'Flight 3: Arrival Date', 'Flight 3: Arrival Time', 'Flight 3: Flight Number',
                'Your HP Office Address - APJ', 'Your HP Office Address - AMS', 'preferredTimezone', 'experience',
                '1st_cnpi', '2nd_cnpi', '3rd_cnpi', 'cnpi_per_tup', 'db_id_df2', 'source_CP', 'db_id'
            ]
            analyzed_reg_df.to_csv("03_final_merged_df.csv", index=False)
            print(f'####################\n03_final_merged_df:')
            analyzed_reg_df.info()
            # 04_roles_df  ----------------------------------------------------------------------------------------    #
            analyzed_reg_df_role = analyzed_reg_df.sort_values(by='role_flag', ascending=True)                          # To sort the DataFrame by the 'Rule_Kept' column (False first, then True)
            cols = [
            'role_flag', 'Registered_or_not', 'Employee Email', 'First name', 'Last name', 'Workstream', 'Role at Show',
            'Tshirt Size', 'Indicate below if you have any special accommodation requests related to your religion or accessibility requirements. Please note that your request is not guaranteed and will be subject for approval.',
            'Comments', 'hc_then_cnpi', 'hc_per_tup', 'cnpi_per_tup', 'hotelsExceptions'
            ]
            analyzed_reg_df_role[cols].to_csv(f"04_roles_df_{today_date}.csv", index=False)
            print(f'####################\n04_roles_df:')
            analyzed_reg_df_role[cols].info()
            # ------------------------------------------------------------------------------------------------------   #
            1/1
            # temp_clean_analyzed_reg_df_cols = [
            # 'Registered_or_not', 'First name', 'Last name', 'Employee Email',
            # 'Workstream', 'Role at Show', 'Drupa role (booth duty) / Visitor / Channel partner',
            # 'Manager name', 'Employee direct manager email', 'Arriving from \nAirport/Country', 'Comments',
            # 'hc_per_tup', 'source_HC', 'hc_then_cnpi', 'source_CP', 'cnpi_per_tup',
            # 'firstName', 'lastName', 'email',
            # 'source_HC', 'hc_per_tup', '1st_hc', '2nd_hc', '3rd_hc',
            # 'source_CP', 'cnpi_per_tup', '1st_cnpi', '2nd_cnpi', '3rd_cnpi',
            # 'phone_number', 'type', 'ticketName', 'status', 'Last seen', 'Registered', 'Last updated date', 'Last updated by', 'Last updated attributes', 'nickName', 'gender',
            # 'Region', 'Country', 'Tshirt Size', 'Will you be attending the sales training day on May 27th?',
            # 'Department/Function', 'Country Coverage', 'Business Segment', 'HP Job Title', 'Manager Name', 'Manager Email Address', 'Type', 'Company Name', 'Job Title',
            # 'Passport Number', 'Passport Expiry Date', 'Passport Issue Date', 'Passport Issued by', 'Date of Birth', 'Dietary Restrictions', 'Life-threatening Allergy', 'Do you require accommodation?',
            # 'Indicate below if you have any special accommodation requests related to your religion or accessibility requirements. Please note that your request is not guaranteed and will be subject for approval.',
            # 'Do you observe Shabbat?', 'Flight 1', 'Flight 2', 'Flight 3', 'Add additional flight? [Flight 2]', 'Add additional flight? [Flight 3]',
            # 'Flight 1: Arrival Airport', 'Flight 1: Arrival Date', 'Flight 1: Arrival Time', 'Flight 1: Flight Number',
            # 'Flight 2: Arrival Airport', 'Flight 2: Arrival Date', 'Flight 2: Arrival Time', 'Flight 2: Flight Number',
            # 'Flight 3: Arrival Airport', 'Flight 3: Arrival Date', 'Flight 3: Arrival Time', 'Flight 3: Flight Number',
            # 'hc_original_index'
            # ]

            # existing_columns = [col for col in temp_clean_analyzed_reg_df_cols if col in analyzed_reg_df.columns]
            # analyzed_reg_df[existing_columns].to_csv(f"04_temp_clean_analyzed_reg_df_{today_date}.csv", index=False)

            # all_analyzed_reg_df_col = [
            #        'Registered_or_not', 'hc_then_cnpi', 'db_id_df1', 'hc_original_index', 'Workstream', 'Drupa role (booth duty) / Visitor / Channel partner',
            #        'Manager name', 'Last name', 'First name', 'Employee Email', 'Employee direct manager email', 'HP Employee#',
            #        'Arriving from \nAirport/Country', 'Comments',
            #        'hc_per_tup', '1st_hc', 'hc_1st_hotel_in', 'hc_1st_hotel_out',
            #        '2nd_hc', 'hc_2nd_hotel_in', 'hc_2nd_hotel_out',
            #        '3rd_hc', 'hc_3rd_hotel_in', 'hc_3rd_hotel_out',
            #        'source_HC',
            #
            #        'firstName', 'lastName', 'jobTitle', 'company', 'country', 'Organization', 'email', 'phone_number', 'type',
            #        'ticketName', 'status', 'OtherSocialMediaURL', 'linkedInProfile', 'telNum', 'twitter', 'hotel', 'hotelCheckIn', 'hotelCheckOut',
            #        'Documents', 'marketingUpdates', 'First seen', 'Login time', 'Last seen', 'Registered', 'checkin',
            #        'Last updated date', 'Last updated by', 'Last updated attributes', 'nickName', 'gender', 'dietary', 'Total watch time',
            #        'Total sessions booked', '_views', 'Region', 'Country', 'APJ - Sub Region & HP Office', 'EMEA - Sub Region & HP Office',
            #        'AMS - Sub Region & HP Office', 'Languages (You may select more than one)', 'If other, please specify language', 'Are you HP employee?',
            #        'HP Business', 'HP Business - Industrial', 'Department/Function', 'Country Coverage', 'Business Segment', 'HP Job Title',
            #        'Role at Show', 'Tshirt Size', 'Manager Name', 'Manager Email Address', 'I am an expert in', 'Type', 'Company Name', 'Job Title',
            #        'Will you be attending the sales training day on May 27th?', 'Terms and Conditions',
            #        'I would like to attend the following dinner events:', 'Additional Info', 'Nationality',
            #        'Passport Number', 'Passport Expiry Date', 'Passport Issue Date', 'Passport Issued by', 'Date of Birth',
            #        'Dietary Restrictions', 'Life-threatening Allergy', 'Do you need any assistance with regards to accessibility?',
            #        'Emergency Contact Name', 'Emergency Contact Number', 'Visa Requirement', 'Do you need a visa for your trip?', 'Accommodation Reminder',
            #        'Do you require accommodation?', 'Reminder',
            #        'Indicate below if you have any special accommodation requests related to your religion or accessibility requirements. Please note that your request is not guaranteed and will be subject for approval.',
            #        '1st Hotel Check in', '1st Hotel Check out',
            #        '2nd Hotel Check in', '2nd Hotel Check out',
            #        '3rd Hotel Check in', '3rd Hotel Check out',
            #        'Do you observe Shabbat?', 'Plan your Trip', 'How will you be travelling to Germany?',
            #        'Flight 1', 'Flight 2', 'Flight 3', 'Add additional flight? [Flight 2]', 'Add additional flight? [Flight 3]',
            #        'Sub Region - APJ', 'Sub Region - AMS', 'Sub Region - EMEA', 'Your HP Office Address - EMEA',
            #        'Flight 1: Arrival Airport', 'Flight 1: Arrival Date', 'Flight 1: Arrival Time', 'Flight 1: Flight Number',
            #        'Flight 2: Arrival Airport', 'Flight 2: Arrival Date', 'Flight 2: Arrival Time', 'Flight 2: Flight Number',
            #        'Flight 3: Arrival Airport', 'Flight 3: Arrival Date', 'Flight 3: Arrival Time', 'Flight 3: Flight Number',
            #        'Your HP Office Address - APJ', 'Your HP Office Address - AMS', 'preferredTimezone', 'experience',
            #        '1st_cnpi', '2nd_cnpi', '3rd_cnpi', 'cnpi_per_tup',
            #        'db_id_df2', 'source_CP', 'db_id']

            # ------------------------------------------------------------------------------------------------------   #

            # Add a new column to indicate if the rule is kept (True) or broken (False)
            # analyzed_reg_df.insert(0, 'role_flag', analyzed_reg_df.apply(lambda row: row['Role at Show'] in Workstream_Pairs_Mapping.get(row['Workstream'], []), axis=1))
            # analyzed_reg_df_role = analyzed_reg_df.sort_values(by='role_flag', ascending=True)                          # To sort the DataFrame by the 'Rule_Kept' column (False first, then True)
            # col_role_df = [
            #     'Employee Email', 'First name', 'Last name', 'role_flag', 'Registered_or_not', 'Workstream', 'Role at Show',
            #     'Tshirt Size', 'Indicate below if you have any special accommodation requests related to your religion or accessibility requirements. Please note that your request is not guaranteed and will be subject for approval.',
            #     'Comments', 'hc_then_cnpi', 'hc_per_tup', 'cnpi_per_tup',
            # ]

            # print(f'####################\n_roles_df:')
            # analyzed_reg_df_role[col_role_df].to_csv(f"__roles_df_{today_date}.csv", index=False)
            # analyzed_reg_df_role[col_role_df].info()
            # print(f'####################\nanalyzed_reg_df:')
            # analyzed_reg_df.to_csv(f"01_temp_analyzed_reg_df_{today_date}.csv", index=False)
            # analyzed_reg_df.info()                                                                                      # df = analyzed_reg_df[analyzed_reg_df['Workstream'].isin(OPEN_WORKSTREAM)].copy()
            1/1
            #   05_hotels_df   ------------------------------------------------------------------------------------    #
            final_df = travels_checks(df=analyzed_reg_df)
            # TODO: excluding non relevance employees (based on their comments) - EXCLUDING_FROM_HOTELS_AND_FLIGHTS_EMAILS
            # final_df = final_df[~final_df['Employee Email'].isin(EXCLUDING_FROM_HOTELS_AND_FLIGHTS_EMAILS)].copy()
            # -----------------------------------------------------------------------------------------------------    #
            cols = [
                'Employee Email', 'First name', 'Last name', 'overall_approval', 'Hotels_inout_approval',
                'Registered_or_not', 'Workstream', 'Role at Show',
                'hc_then_cnpi', '1st_in', '1st_out', '2nd_in', '2nd_out', '3rd_in', '3rd_out',
                'type', 'ticketName', 'gender', 'Country', 'Passport Expiry Date',
                'Tshirt Size',
                'Indicate below if you have any special accommodation requests related to your religion or accessibility requirements. Please note that your request is not guaranteed and will be subject for approval.',
                'hc_per_tup', 'cnpi_per_tup', 'hc_original_index', 'hotelsExceptions'
            ]
            final_df[cols].to_csv(f"05_hotels_df_{today_date}.csv", index=False)
            print(f'####################\n05_hotels_df:')
            final_df[cols].info()
            # 06_discrepancies_df   -------------------------------------------------------------------------------    #
            final_df[COLS_FOR_LIMOR].to_excel(f"06_discrepancies_df_{today_date}.xlsx", index=False, engine='openpyxl')     # final_df[COLS_FOR_LIMOR].to_csv(f"__discrepancies_{today_date}.csv", index=False)
            print(f'####################\n06_discrepancies_df_:')
            final_df[COLS_FOR_LIMOR].info()
            if not DEV_MODE:
                os.chdir(LIMOR_FOLDER)                                                                                  # Change the current working directory to target folder
                final_df[COLS_FOR_LIMOR].to_excel(f"__discrepancies_{today_date}.xlsx", index=False, engine='openpyxl') # final_df[COLS_FOR_LIMOR].to_csv(f"__discrepancies_{today_date}.csv", index=False)
                print(f'####################\n06_discrepancies_df:')
                # TODO: My Excel editor faunction
                # excel_file_editor(filePath='', fileName=f"__discrepancies_{today_date}.xlsx")
                os.chdir(ORG_CWD)
            # 07 Export status per workstream  --------------------------------------------------------------------    #
            export_discrepancies_per_workstream(fin_df=final_df)
            #   ---------------------------------------------------------------------------------------------------    #
            self.fin_df = final_df.copy()

            print(self.fin_df.columns.to_list())

            self.tab.after(0, lambda: self.update_gui_after_processing(message="Done!"))
        except Exception as e:
            error_message = f"Failed to analysing due to:\n{e}"
            print(error_message)
            self.tab.after(0, lambda: messagebox.showerror(f"Error2 - {identifier}", error_message))
    #   TODO: ---------------------------------------------------------------------------------------------------      #
    #   TODO: ---------------------------------------------------------------------------------------------------      #
    #   TODO: ---------------------------------------------------------------------------------------------------      #
    def flights_analysis(self, identifier, post_to_gui):
        self.last_selected_button = identifier
        flights_report_file_name = f"__flights_{today_date}.xlsx"
        cols_for_limor = [
            'Employee Email', 'First name', 'Last name', 'overall_approval', 'Registered_or_not',
            'Hotels_inout_approval', 'role_flag', 'Workstream', 'Role at Show',
            'Drupa role (booth duty) / Visitor / Channel partner',
            'hc_then_cnpi', '1st_in', '1st_out', '2nd_in', '2nd_out', '3rd_in', '3rd_out',
            'Comments', 'type', 'ticketName', 'gender', 'Country', 'Passport Expiry Date', 'Tshirt Size',
            'Indicate below if you have any special accommodation requests related to your religion or accessibility requirements. Please note that your request is not guaranteed and will be subject for approval.',
            'hc_per_tup', 'cnpi_per_tup', '1st Hotel Check in', '1st Hotel Check out', '2nd Hotel Check in',
            '2nd Hotel Check out', '3rd Hotel Check in', '3rd Hotel Check out', 'hc_original_index']
        columns_for_efrat = [
                            'empID', 'PNR', '#Trips', 'workstream', 'email', 'source',
                            'First name', ' Last name', 'Organization',
                            'Total drupa days', 'Total flights cost',
                            'Other total cost per day', 'Cost Center'
                            ]
        def dfs_info(df):
            print(df.info())
            print(df.columns.to_list())
            pass
        # def convert_cwt_dates(date_series):
        #     current_year = pd.to_datetime('today').year
        #     return pd.to_datetime(date_series + '-' + str(current_year), format='%d-%b-%Y')                             # Convert to datetime with format "d-MMM" and add current year
        def convert_cwt_dates(date_series):
            converted_dates = pd.Series(index=date_series.index, dtype='datetime64[ns]')                                # Initialize the converted_dates Series with datetime type
            current_year = pd.to_datetime('today').year                                                                 # Define the current year
            for idx, date in date_series.iteritems():                                                                   # Loop through the date series and convert according to the format
                if isinstance(date, str):
                    try:        # Try converting the "dd-MMM" format
                        date += f'-{current_year}'
                        converted_date = pd.to_datetime(date, format='%d-%b-%Y')
                    except:
                        try:    # Try converting the "dd/MM/yyyy" format
                            converted_date = pd.to_datetime(date, format='%d/%m/%Y')
                        except:
                            converted_date = pd.NaT
                else:
                    converted_date = pd.to_datetime(date, errors='coerce')                                              # If it's not a string, it might already be in datetime format or None/NaN
                converted_dates.at[idx] = converted_date
            return converted_dates
        def cwt_trips_preprocess(df):
            # cols=[
            #     'Traveler Name', 'Employee #',
            #     'Hierarchy 1', 'Hierarchy 2',
            #     'Cost Location', 'Business Unit',
            #     'O&D Airport Pair', 'PNR', 'Validating Airline Name',
            #     'Ticket Departure Date', 'Ticket Return Date',
            #     'Transactions', 'Spend',
            #     'Transactions.1', 'Spend.1',
            #     'Transactions.2', 'Spend.2',
            #     'Transactions.3', 'Spend.3'
            #     ]
            df.to_csv(f"step_00.csv", index=False)
            # Convert the start and end dates to datetime if they aren't already
            DRUPA_DATES['start'] = pd.to_datetime(DRUPA_DATES['start'])
            DRUPA_DATES['end'] = pd.to_datetime(DRUPA_DATES['end'])
            # Check and convert 'Departure Date'
            if 'Ticket Departure Date' in df.columns:
                df.loc[:, 'arrivalDate'] = convert_cwt_dates(df['Ticket Departure Date'])                               # df['arrivalDate'] = convert_cwt_dates(df['Ticket Departure Date'])
            elif 'Departure Date' in df.columns:
                df.loc[:, 'arrivalDate'] = convert_cwt_dates(df['Departure Date'])                                      # df['arrivalDate'] = convert_cwt_dates(df['Departure Date'])
            # Check and convert 'Return Date'
            if 'Ticket Return Date' in df.columns:
                df.loc[:, 'returnDate'] = convert_cwt_dates(df['Ticket Return Date'])                                   # df['returnDate'] = convert_cwt_dates(df['Ticket Return Date'])
            elif 'Return Date' in df.columns:
                df.loc[:, 'returnDate'] = convert_cwt_dates(df['Return Date'])                                          # df['returnDate'] = convert_cwt_dates(df['Return Date'])
            # First filter: keep rows where column A is within the start and end date range
            df = df[(df['arrivalDate'] >= DRUPA_DATES['start']) & (df['arrivalDate'] <= DRUPA_DATES['end'])].copy()     # df = df[(df['arrivalDate'] >= DRUPA_DATES['start']) & (df['arrivalDate'] <= DRUPA_DATES['end'])]            df.to_csv(f"aaaaaa.csv", index=False)   # TODO: DROPPPP

            valid_empIDs = self.hc_comb_df['EmpNum'].unique().tolist()
            valid_empIDs = [empID.strip("'") for empID in valid_empIDs]                                                 # Remove leading single quotes
            print(f"Unique IDs: {valid_empIDs}")
            print(f"emp IDs: {df['Employee #'].unique().tolist()}")

            # Second filter: keep rows where the ID column values are in the valid_ids list
            df.loc[:, 'Employee #']                = df['Employee #'].astype(str).str.replace('\D', '', regex=True)
            df                                     = df[df['Employee #'].isin(valid_empIDs)]
            df.loc[:, 'overlay']                   = False
            split_names = df['Traveler Name'].fillna('').astype(str).str.split('/', n=1, expand=True)
            # If after splitting, any row doesn't have two elements, fill the missing element with an empty string
            split_names[0] = split_names[0].fillna('')
            split_names[1] = split_names[1].fillna('')
            # Now assign the split names to new columns in df
            df['LastName'] = split_names[0]
            df['FirstName'] = split_names[1]
            # df[['LastName', 'FirstName']]   = df['Traveler Name'].str.split('/', 1, expand=True)
            # df['FirstName']                 = df['FirstName'].str.replace(' MR', '', regex=False).str.strip()
            df.loc[:, 'Days Between']              = (df['returnDate'] - df['arrivalDate']).dt.days                            # Calculate the difference in days
            spend_columns                   = [col for col in df.columns.to_list() if col.lower().startswith('spend')]  # Filter out columns that start with 'Spend' and convert 'spend' to 'Spend' to match the pattern
            spend_columns_sorted            = sorted(spend_columns, key=lambda x: (int(x.split('.')[1]) if '.' in x else 0))
            last_spend_column               = spend_columns_sorted[-1]                                                  # The last spend is the spend with the highest x value spend_x, spend_x+1..
            # df.rename(columns={last_spend_column : 'Total legs of trips'})
            def adjust_days_between(group):
                """ Adjust 'Days Between' for duplicated trips"""
                # if len(group) > 1 and group['Ticket Departure Date'].nunique() == 1 and group['Ticket Return Date'].nunique() == 1:
                if len(group) > 1 and group['arrivalDate'].nunique() == 1 and group['returnDate'].nunique() == 1:
                    group['Original Days Between'] = group['Days Between']
                    group['Days Between'] = group['Days Between'] / len(group)
                    group['overlay'] = True                                                                             # Flag this group as having an overlap
                return group
            df = df.groupby(['PNR', 'Employee #'], group_keys=False).apply(adjust_days_between)                         # df = df.groupby(['PNR', 'Employee #']).apply(adjust_days_between).reset_index(drop=True)
            cols = [
            'Employee #', 'Traveler Name', 'arrivalDate', 'returnDate', 'Days Between', 'PNR', f'{last_spend_column}', 'overlay'
            ]
            mode = 'a' if os.path.exists(flights_report_file_name) else 'w'                                             # Use ExcelWriter in append mode to add to an existing file without overwriting
            with pd.ExcelWriter(flights_report_file_name, engine='openpyxl', mode=mode) as writer:
                df[cols].to_excel(writer, sheet_name="step1", index=False)
            df.to_csv(f"step1.csv", index=False)
            return df
        def agg_def(df):
            """
            :param df:
            :return:
            """
            aggregated_data = pd.DataFrame()                                                                            # Create an empty DataFrame to store the aggregated data
            # Group by both PNR and Employee #, and iterate over each group
            for (pnr, employee_number), group_data in df.groupby(['PNR', 'Employee #']):
                # Create a dictionary to hold the sum of each numeric column for this group
                agg_data = {col: group_data[col].sum() for col in group_data.columns if pd.api.types.is_numeric_dtype(group_data[col])}
                # Add the PNR, Employee #, and the count of occurrences to the dictionary
                agg_data['PNR'] = pnr
                agg_data['Employee #'] = employee_number
                agg_data['Occurrences'] = len(group_data)
                # Aggregate non-numeric columns by concatenating unique string representations
                for col in group_data.columns:
                    if not pd.api.types.is_numeric_dtype(group_data[col]) and col not in ['PNR', 'Employee #']:
                        agg_data[col] = ', '.join(group_data[col].astype(str).unique())
                # Append to the aggregated data
                aggregated_data = pd.concat([aggregated_data, pd.DataFrame([agg_data])], ignore_index=True)
            # Rearrange the columns to have PNR, Employee #, and Occurrences at the front
            column_order = ['PNR', 'Employee #', 'Occurrences'] + [col for col in aggregated_data.columns if col not in ['PNR', 'Employee #', 'Occurrences']]
            aggregated_data = aggregated_data[column_order]
            aggregated_data['overlay'] = aggregated_data['overlay']>0

            # Use ExcelWriter in append mode to add to an existing file without overwriting
            aggregated_data.to_csv("step2.csv", index=False)
            mode = 'a' if os.path.exists(flights_report_file_name) else 'w'
            with pd.ExcelWriter(flights_report_file_name, engine='openpyxl', mode=mode) as writer:
                aggregated_data.to_excel(writer, sheet_name="step2", index=False)
            return aggregated_data
        def agg_per_emp(df):
            # Define aggregation rules: non-numeric columns will be aggregated into lists, while numeric columns will be summed.
            agg_rules = {col: 'sum' if pd.api.types.is_numeric_dtype(df[col]) else (lambda x: list(x)) for col in df.columns if col != 'Employee #'}
            # Adjusting the aggregation rules for specific columns to avoid summing up non-related numeric identifiers or similar
            non_sum_columns = ['PNR', 'Cost Location', 'Business Unit', 'O&D Airport Pair', 'Validating Airline Name',
                               'Ticket Departure Date', 'Ticket Return Date', 'arrivalDate', 'returnDate', 'LastName',
                               'FirstName']
            for col in non_sum_columns:
                if col in agg_rules:
                    agg_rules[col] = lambda x: list(x)
            # Perform aggregation
            aggregated_df = df.groupby('Employee #').agg(agg_rules)
            # Creating "Tot_{column}" for each numeric column
            for col in df.select_dtypes(include='number').columns:
                if col != 'Employee #':
                    aggregated_df[f'Tot_{col}'] = aggregated_df[col]
            # Since we're summing up the numeric columns twice (once in the original aggregation and then explicitly for Tot_columns),
            # we'll remove the original numeric columns from the aggregated dataframe
            aggregated_df = aggregated_df.drop(columns=df.select_dtypes(include='number').columns)
            aggregated_df.reset_index(inplace=True)
            aggregated_df['overlay'] = aggregated_df['overlay'] > 0
            aggregated_df.to_csv("step3.csv", index=False)
            mode = 'a' if os.path.exists(flights_report_file_name) else 'w'
            with pd.ExcelWriter(flights_report_file_name, engine='openpyxl', mode=mode) as writer:
                aggregated_df.to_excel(writer, sheet_name="step3", index=False)
            return aggregated_df
        def preparation_for_exportation(df):
            print(df.columns.to_list())

            # Use regex to find columns that match the pattern 'Tot_Spend.X'
            pattern = re.compile(r'Tot_Transactions\.\d+')
            matching_columns = [col for col in df.columns if pattern.match(col)]
            # Extract the number part and find the column with the maximum number
            last_column = max(matching_columns, key=lambda x: int(x.split('.')[-1]))

            df.rename(columns={'Employee #': 'ID',
                               'Cost Location': 'Cost center',
                               'Cost location': 'Cost center',
                               'Tot_Days Between': 'Total drupa days',
                                last_column : 'Total legs of trips',
                               'Tot_AggTripCost': 'Total trips costs'},
                                inplace=True)
            df['Total trips costs'] = df['Total trips costs'].astype(int)
            print(df.columns.to_list())

            df['Total trips'] = df['PNR'].apply(len)
            df['HC_flag'] = df['ID'].isin(self.hc_comb_df['HP Employee#'])
            df['FirstName']=df['FirstName'].apply(lambda x: x[0] if x else '')
            df['LastName']=df['LastName'].apply(lambda x: x[0] if x else '')
            df['Organization']=df['Organization'].apply(lambda x: x[0] if x else '')
            df['Cost center']=df['Cost center'].apply(lambda x: x[0] if x else '')
            return_cols = [
                'ID', 'FirstName', 'LastName', 'Organization', 'Cost center', 'Total drupa days',
                'Total trips costs', 'Total trips', 'Total legs of trips',
                'PNR', 'HC_flag'
            ]
            df[return_cols].to_csv("step4_final_result.csv", index=False)
            df[return_cols].to_csv(f"__flights_status_{today_date}.csv", index=False)

            mode = 'a' if os.path.exists(flights_report_file_name) else 'w'
            with pd.ExcelWriter(flights_report_file_name, engine='openpyxl', mode=mode) as writer:
                df[return_cols].to_excel(writer, sheet_name="final_result", index=False)
            return
        def ofir_filter_trips_per_empNum():
            cols=['Pfile no', 'customer name', 'from', 'to', 'Dep.Time', 'Arr.Time', 'Airline', 'class', 'Airfare',
                  'Tax', 'Total', 'amount saving', 'reason', 'CCLIENT_NAME',
                  'CUSTOMER_ID', 'area', 'pricelevelclassAGENT_REF', 'PNR']
            pass
        # prev_df = self.fin_df.copy()
        dfs_info(df=self.cwt_df)    # dfs_info(df=self.ofir_df)
        # TODO:
        temp_df = cwt_trips_preprocess(df=self.cwt_df)
        temp_df = agg_def(df=temp_df)
        temp_df.rename(columns={'Spend.3': 'AggTripCost', 'Occurrences': 'LegsOfTrip', 'Hierarchy 2': 'Organization'}, inplace=True)
        temp_df = agg_per_emp(df=temp_df)
        preparation_for_exportation(df=temp_df)
        print("Done")
    #   TODO: ---------------------------------------------------------------------------------------------------      #
    #   TODO: ---------------------------------------------------------------------------------------------------      #
    #   TODO: ---------------------------------------------------------------------------------------------------      #
    def employees_extractor(self, identifier, post_to_gui):
        self.last_selected_button = identifier
        emp_data = []
        import tkinter.simpledialog as simpledialog
        import webbrowser
        import pyautogui
        from PIL import Image, ImageEnhance, ImageFilter
        import pytesseract
        import re
        def get_user_name():
            if OCR_FLAG==1:
                for index, row in self.hc_comb_df.iterrows():
                    # user_name = simpledialog.askstring("Input", "What is your name?")
                    user_name = row['Employee Email']
                    print(f"\n{user_name}:")
                    if user_name:
                        try:
                            url = f"https://directoryworks.hpicorp.net/protected/people/view/person/ldif/?dn=uid%3D{user_name.split('@')[0]}%40hp.com%2Cou%3DPeople%2Co%3Dhp.com"
                            webbrowser.open(url)
                            time.sleep(5)
                            pyautogui.screenshot('before_click.png', region=(0, 400, 1000, 1000))
                            time.sleep(3)
                            image_path = 'before_click.png'
                            img = Image.open(image_path)

                            img = img.convert('L')  # Convert to grayscale
                            img = img.filter(ImageFilter.MedianFilter())  # Apply a median filter
                            enhancer = ImageEnhance.Contrast(img)
                            img = enhancer.enhance(2)  # Enhance contrast

                            pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
                            custom_config = r'--oem 3 --psm 6 -c tessedit_char_whitelist=abcdefghijklmnopqrstuvwxyz0123456789'
                            text = pytesseract.image_to_string(img, config=custom_config)
                            match = re.search(r"employeenumbers*(\d+)", text, re.IGNORECASE)

                            print(f"{user_name}: {match.group(1)}")
                            emp_record = {
                                "name": user_name.split('@')[0],
                                "id": match.group(1) if match else "Not Found",
                                "email": user_name,
                                "hc_index": index
                                }
                            emp_data.append(emp_record)
                        except Exception as e:
                            print(e)
                    else:
                        print("No name was provided.")
                        # Use the passed `post_to_gui` function to safely execute `get_user_name` on the GUI thread
                emp_df = pd.DataFrame(emp_data)
                emp_df.to_csv('emp_ocr.csv', index=False)
            elif OCR_FLAG==2:
                x=520
                for index, row in self.hc_comb_df.iloc[x:].iterrows():
                    # user_name = simpledialog.askstring("Input", "What is your name?")
                    user_name = row['Employee Email']
                    print(f"\n{user_name}:")
                    if user_name:
                        try:
                            url = f"https://directoryworks.hpicorp.net/protected/people/view/person/ldif/?dn=uid%3D{user_name.split('@')[0]}%40hp.com%2Cou%3DPeople%2Co%3Dhp.com"
                            webbrowser.open(url)
                            print(f"{index}: {user_name}")
                            time.sleep(5)
                        except Exception as e:
                            print(e)
            elif OCR_FLAG==3:
                with open('all.txt', 'r') as file:
                    text = file.read()
                # Extracting email and employee numbers using regex
                pattern = re.compile(r"(\S+@hp.com):\s*(\d+)")
                matches = pattern.findall(text)
                # Converting to DataFrame
                emp_df = pd.DataFrame(matches, columns=["email", "empNum"])
                emp_df.to_csv('emp.csv', index=False)
            else:
                user_name = simpledialog.askstring("Input", "What is your name?")
                if user_name:
                    try:
                        url = f"https://directoryworks.hpicorp.net/protected/people/view/person/ldif/?dn=uid%3D{user_name.split('@')[0]}%40hp.com%2Cou%3DPeople%2Co%3Dhp.com"
                        webbrowser.open(url)
                    except Exception as e:
                        print(e)
        post_to_gui(get_user_name)
    #   TODO: ---------------------------------------------------------------------------------------------------      #
    #   TODO: ---------------------------------------------------------------------------------------------------      #
    #   TODO: ---------------------------------------------------------------------------------------------------      #
#   -------------------------------------------------------------------------------------------------------------      #
class MultiTabGUI:
    def __init__(self, master):
        self.master = master
        self.master.title(f"Drupa Analysis App ({VER[0]})")
        high = 600
        self.master.geometry(f"{2*high}x{high}")
        # Create the tab control
        self.tab_control = ttk.Notebook(self.master)
        # Create the first tab  ------------------------------------------------------------------------------------   #
        self.TAB1 = ttk.Frame(self.tab_control)
        self.tab_control.add(self.TAB1, text='ExcelFileSelector')
        self.excel_selector = ExcelFileSelector(self.TAB1)                                                              # Instantiate the ExcelFileSelector in the first tab
        #   --------------------------------------------------------------------------------------------------------   #
        # Create the second tab  -----------------------------------------------------------------------------------   #
        self.TAB2 = ttk.Frame(self.tab_control)
        self.tab_control.add(self.TAB2, text='Second Tab')
        self.label2 = tk.Label(self.TAB2, text="This is the second tab")                                                # Add widgets to the second tab
        self.label2.pack()
        #   --------------------------------------------------------------------------------------------------------   #
        # Pack the tab control to make the tabs visible
        self.tab_control.pack(expand=1, fill="both")
        #   --------------------------------------------------------------------------------------------------------   #
#   -------------------------------------------------------------------------------------------------------------      #
def create_gui():
    root = tk.Tk()
    app = MultiTabGUI(root)                                                                                             # app = ExcelFileSelector(root)
    root.mainloop()
#   -------------------------------------------------------------------------------------------------------------      #
#   -------------------------------------------------------------------------------------------------------------      #
#   -------------------------------------------------------------------------------------------------------------      #
if __name__ == '__main__':
    create_gui()
#   -------------------------------------------------------------------------------------------------------------      #
#   -------------------------------------------------------------------------------------------------------------      #
#   -------------------------------------------------------------------------------------------------------------      #
#   -------------------------------------------------------------------------------------------------------------      #
#   -------------------------------------------------------------------------------------------------------------      #
#   -------------------------------------------------------------------------------------------------------------      #
#   -------------------------------------------------------------------------------------------------------------      #
#   -------------------------------------------------------------------------------------------------------------      #
#   -------------------------------------------------------------------------------------------------------------      #
#   -------------------------------------------------------------------------------------------------------------      #
#   -------------------------------------------------------------------------------------------------------------      #
#   -------------------------------------------------------------------------------------------------------------      #
#   -------------------------------------------------------------------------------------------------------------      #
#   -------------------------------------------------------------------------------------------------------------      #
#   -------------------------------------------------------------------------------------------------------------      #
#   -------------------------------------------------------------------------------------------------------------      #
#   -------------------------------------------------------------------------------------------------------------      #
#   -------------------------------------------------------------------------------------------------------------      #
#   -------------------------------------------------------------------------------------------------------------      #
#   -------------------------------------------------------------------------------------------------------------      #
#   -------------------------------------------------------------------------------------------------------------      #
#   -------------------------------------------------------------------------------------------------------------      #
#   -------------------------------------------------------------------------------------------------------------      #
#   -------------------------------------------------------------------------------------------------------------      #